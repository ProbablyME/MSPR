"""
extract_flights.py
==================
Pipeline ETL UNIQUE — données de vols Eurocontrol (OPDI) + données CO2 EASA.

Architecture en 3 phases, tout en mémoire (aucun fichier intermédiaire) :

  1. EXTRACT
       a) Lecture du fichier Excel EASA (co2 par type d'avion)
       b) Chargement de l'index aéroports depuis airports.csv
       c) Téléchargement des fichiers .parquet mensuels Eurocontrol
          directement dans un buffer mémoire (aucun fichier disque)

  2. TRANSFORM
       a) Agrégation CO2 EASA par désignation de type
       b) Filtrage des vols (adep ET ades non nuls/vides)
          Enrichissement coordonnées + distance Haversine
          Comptage typecodes ICAO pour trouver le dominant par route

  3. LOAD  (dans cet ordre pour permettre le calcul CO2)
       a) mart.dim_vehicle_avion  (données EASA — migration idempotente)
       b) mart.dim_station        (aéroports avec lat, lon, nom)
       c) mart.dim_route_avion    (routes avec distance_km + dominant_typecode)
          -> UPDATE co2_total_kg = distance_km × co2_per_km via JOIN EASA

Résumé IA :
  - Checkpoints : sauvegarde après chaque mois traité → reprise automatique si crash
  - Try/except  : chaque étape est isolée, les erreurs n'arrêtent pas le pipeline

Workflow : docker compose up -d  →  python extract_flights.py
"""

import io
import os
import sys
import pickle
import logging
from math import radians, sin, cos, asin, sqrt
from datetime import datetime
from statistics import mean
from collections import defaultdict
from dateutil.relativedelta import relativedelta

import openpyxl
import pyarrow.parquet as pq
import pyarrow.compute as pc
import pyarrow.csv as pa_csv
import requests
import psycopg2
from psycopg2.extras import execute_values

# ── Logging ───────────────────────────────────────────────────────────────────
LOG_FILE = f"extract_flights_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════
AIRPORTS_CSV     = "./airports.csv"
EASA_FILE        = "./EASA_CO2_Database__web__-Issue11.xlsx"
EASA_SHEET       = "EASA CO2DB (web)"
CHECKPOINT_FILE  = "./etl_checkpoint.pkl"

BASE_URL   = "https://www.eurocontrol.int/performance/data/download/OPDI/v002/flight_list/flight_list_"
START_DATE = "202501"
END_DATE   = "202512"

DB_CONFIG = {
    "host":     os.environ.get("DB_HOST",     "localhost"),
    "port":     os.environ.get("DB_PORT",     "5433"),
    "database": os.environ.get("DB_NAME",     "postgres"),
    "user":     os.environ.get("DB_USER",     "postgres"),
    "password": os.environ.get("DB_PASSWORD", "lpironti"),
}

# Mapping ICAO prefix → code ISO pays
ICAO_PREFIX_TO_ISO: dict[str, str] = {
    "LF": "FR", "LI": "IT", "LE": "ES", "LP": "PT",
    "LH": "HU", "LK": "CZ", "LZ": "SK", "LO": "AT",
    "LJ": "SI", "LD": "HR", "LQ": "BA", "LY": "RS",
    "LG": "GR", "LB": "BG", "LR": "RO", "LU": "MD",
    "LT": "TR", "LM": "MT", "LC": "CY", "LA": "AL",
    "LW": "MK", "LX": "GI", "LN": "MC",
    "EG": "GB", "ED": "DE", "ET": "DE", "EH": "NL",
    "EB": "BE", "LS": "CH", "EK": "DK", "EN": "NO",
    "EF": "FI", "ES": "SE", "EE": "EE", "EV": "LV",
    "EY": "LT", "EI": "IE", "EL": "LU", "EP": "PL",
    "UM": "BY", "UK": "UA", "UL": "RU", "UU": "RU",
}

# Correspondance désignation EASA → code ICAO type (Source : ICAO Doc 8643)
EASA_TO_ICAO: dict[str, str] = {
    # Airbus A319neo
    "A319-151N":  "A19N",
    "A319-153N":  "A19N",
    "A319-171N":  "A19N",
    # Airbus A320neo (moteurs CFM LEAP)
    "A320-251N":  "A20N",
    "A320-252N":  "A20N",
    "A320-253N":  "A20N",
    # Airbus A320neo (moteurs Pratt & Whitney PW1100G)
    "A320-271N":  "A20N",
    "A320-272N":  "A20N",
    "A320-273N":  "A20N",
    # Airbus A321neo / XLR
    "A321-251NX": "A21N",
    "A321-252NX": "A21N",
    "A321-253NX": "A21N",
    "A321-253NY": "A21N",
    "A321-271NX": "A21N",
    "A321-271NY": "A21N",
    "A321-272NX": "A21N",
    # Airbus A330neo
    "A330-841":   "A338",
    "A330-941":   "A339",
    # Airbus A350
    "A350-941":   "A359",
    "A350-1041":  "A35K",
    # ATR
    "ATR 72-212A": "AT72",
    # Airbus Canada A220 (ex Bombardier C-Series)
    "BD-500-1A10": "BCS1",   # A220-100
    "BD-500-1A11": "BCS3",   # A220-300
    # Dassault
    "Falcon 7X":  "F7X",
}


# ═══════════════════════════════════════════════════════════════════════════════
# CHECKPOINT
# ═══════════════════════════════════════════════════════════════════════════════

def save_checkpoint(
    completed_months: list[str],
    all_airports: dict,
    all_routes: dict,
    all_tc_counts: dict,
) -> None:
    """
    Sauvegarde l'état de progression sur disque (pickle).
    Appelé après chaque mois traité avec succès.
    Si la sauvegarde échoue, un warning est loggé mais le pipeline continue.
    """
    try:
        state = {
            "completed_months": completed_months,
            "all_airports":     all_airports,
            "all_routes":       all_routes,
            "all_tc_counts":    all_tc_counts,
        }
        with open(CHECKPOINT_FILE, "wb") as f:
            pickle.dump(state, f)
        logger.info(
            f"[CHECKPOINT] Sauvegarde OK — {len(completed_months)} mois traités."
        )
    except Exception as e:
        logger.warning(f"[CHECKPOINT] Echec de la sauvegarde : {e} — pipeline continue.")


def load_checkpoint() -> dict | None:
    """
    Charge l'état de progression depuis le disque.
    Retourne None si le fichier n'existe pas ou est corrompu.
    """
    if not os.path.exists(CHECKPOINT_FILE):
        return None
    try:
        with open(CHECKPOINT_FILE, "rb") as f:
            state = pickle.load(f)
        done = state.get("completed_months", [])
        logger.info(
            f"[CHECKPOINT] Reprise détectée — {len(done)} mois déjà traités : "
            f"{', '.join(done)}"
        )
        return state
    except Exception as e:
        logger.warning(
            f"[CHECKPOINT] Fichier illisible ({e}) — redémarrage complet."
        )
        return None


def clear_checkpoint() -> None:
    """Supprime le fichier de checkpoint après une fin de pipeline réussie."""
    try:
        if os.path.exists(CHECKPOINT_FILE):
            os.remove(CHECKPOINT_FILE)
            logger.info("[CHECKPOINT] Supprimé — pipeline terminé avec succès.")
    except Exception as e:
        logger.warning(f"[CHECKPOINT] Impossible de supprimer le fichier : {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _get_country(icao: str) -> str:
    if icao and len(icao) >= 2:
        return ICAO_PREFIX_TO_ISO.get(icao[:2].upper(), "XX")
    return "XX"


def _haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Distance orthodromique entre deux points (km)."""
    R = 6_371.0
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    a = sin((lat2 - lat1) / 2) ** 2 + cos(lat1) * cos(lat2) * sin((lon2 - lon1) / 2) ** 2
    return 2 * R * asin(sqrt(a))


def _classify_service(mtom_kg: int | None) -> str:
    """
    Classe un avion selon sa MTOM :
      < 40 t   -> regional
      40-100 t -> court_moyen_courrier
      > 100 t  -> long_courrier
    """
    if mtom_kg is None:
        return "inconnu"
    if mtom_kg < 40_000:
        return "regional"
    if mtom_kg < 100_000:
        return "court_moyen_courrier"
    return "long_courrier"


def _generate_urls() -> list[str]:
    start_dt = datetime.strptime(START_DATE, "%Y%m")
    end_dt   = datetime.strptime(END_DATE,   "%Y%m")
    urls, current = [], start_dt
    while current <= end_dt:
        urls.append(f"{BASE_URL}{current.strftime('%Y%m')}.parquet")
        current += relativedelta(months=1)
    return urls


# Type : (lat, lon, name, country_code)
AirportInfo = tuple[float, float, str, str]
# Type : (distance_km | None, dominant_typecode | None)
RouteInfo = tuple[float | None, str | None]


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 1 — EXTRACT
# ═══════════════════════════════════════════════════════════════════════════════

def extract_easa(filepath: str, sheet: str) -> dict | None:
    """
    Lit la feuille EASA CO2DB et agrège les valeurs CO2 par désignation de type.

    Colonnes utilisées (index 0-based) :
      [2]  Aeroplane Type Certificate Holder  -> manufacturer
      [3]  Aeroplane Type Designation         -> label / clé
      [8]  Aeroplane MTOM (kg)                -> mtom_kg
      [14] Number of Engines                  -> num_engines
      [26] CO2 Metric Value (kg/km)           -> co2_per_km

    Retourne None si le fichier est introuvable ou illisible.
    """
    logger.info(f"[EXTRACT] EASA : {filepath} (feuille : {sheet})")
    try:
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Fichier EASA introuvable : {filepath}")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        if sheet not in wb.sheetnames:
            raise ValueError(
                f"Feuille '{sheet}' absente. Feuilles disponibles : {wb.sheetnames}"
            )

        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[3:]  # Les 3 premières lignes sont des en-têtes de section

        types: dict = defaultdict(
            lambda: {"co2_values": [], "mtom": None, "engines": None, "holder": None}
        )
        for r in data_rows:
            designation = r[3]
            co2         = r[26]
            if not designation or co2 is None:
                continue
            key = str(designation).strip()
            try:
                types[key]["co2_values"].append(float(co2))
                if types[key]["mtom"]    is None and r[8]  is not None:
                    types[key]["mtom"]    = int(r[8])
                if types[key]["engines"] is None and r[14] is not None:
                    types[key]["engines"] = int(r[14])
                if types[key]["holder"]  is None and r[2]  is not None:
                    types[key]["holder"]  = str(r[2]).strip()
            except (ValueError, TypeError) as e:
                logger.warning(f"  -> Ligne EASA ignorée ({key}) : {e}")

        if not types:
            raise ValueError("Aucun type d'avion extrait — vérifiez la structure du fichier.")

        logger.info(f"  -> {len(types)} types d'avions distincts extraits.")
        return dict(types)

    except FileNotFoundError as e:
        logger.error(f"[EXTRACT] EASA — fichier manquant : {e}")
    except ValueError as e:
        logger.error(f"[EXTRACT] EASA — données invalides : {e}")
    except Exception as e:
        logger.error(f"[EXTRACT] EASA — erreur inattendue : {e}")
    return None


def load_airports_index(csv_path: str) -> dict[str, AirportInfo]:
    """
    Construit un dictionnaire de lookup depuis airports.csv.
    Clé : code ICAO (colonne 'ident'), valeur : (lat, lon, name, country).
    Retourne un dict vide si le fichier est introuvable.
    """
    logger.info(f"[EXTRACT] Aéroports : {csv_path}")
    try:
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"Fichier aéroports introuvable : {csv_path}")

        table = pa_csv.read_csv(csv_path)

        required = {"ident", "latitude", "longitude", "name", "country"}
        missing = required - set(table.schema.names)
        if missing:
            raise ValueError(f"Colonnes manquantes dans airports.csv : {missing}")

        index: dict[str, AirportInfo] = {}
        idents    = table.column("ident").to_pylist()
        lats      = table.column("latitude").to_pylist()
        lons      = table.column("longitude").to_pylist()
        names     = table.column("name").to_pylist()
        countries = table.column("country").to_pylist()

        for ident, lat, lon, name, country in zip(idents, lats, lons, names, countries):
            try:
                if ident and lat is not None and lon is not None:
                    index[str(ident).strip()] = (
                        float(lat), float(lon), str(name), str(country)
                    )
            except (ValueError, TypeError) as e:
                logger.warning(f"  -> Aéroport ignoré ({ident}) : {e}")

        logger.info(f"  -> {len(index):,} aéroports indexés.")
        return index

    except FileNotFoundError as e:
        logger.error(f"[EXTRACT] Aéroports — fichier manquant : {e}")
    except ValueError as e:
        logger.error(f"[EXTRACT] Aéroports — données invalides : {e}")
    except Exception as e:
        logger.error(f"[EXTRACT] Aéroports — erreur inattendue : {e}")
    return {}


def extract_parquet(url: str):
    """
    Télécharge le fichier .parquet depuis l'URL dans un buffer mémoire.
    Aucun fichier n'est écrit sur disque.
    Retourne None en cas d'échec (HTTP, réseau, parsing).
    """
    logger.info(f"[EXTRACT] Vols : {url.split('/')[-1]}")
    try:
        response = requests.get(url, timeout=180)
        response.raise_for_status()
        size_mb = len(response.content) / 1_000_000
        table = pq.read_table(io.BytesIO(response.content))
        logger.info(f"  -> {len(table):,} lignes  ({size_mb:.1f} Mo)")
        return table
    except requests.exceptions.HTTPError as e:
        logger.warning(f"  -> HTTP {e.response.status_code} — fichier ignoré.")
    except requests.exceptions.ConnectionError:
        logger.warning("  -> Erreur de connexion — fichier ignoré.")
    except requests.exceptions.Timeout:
        logger.warning("  -> Timeout (180 s dépassé) — fichier ignoré.")
    except Exception as e:
        logger.warning(f"  -> Erreur inattendue lors du téléchargement : {e}")
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 2 — TRANSFORM
# ═══════════════════════════════════════════════════════════════════════════════

def transform_easa(types: dict) -> list[tuple]:
    """
    Construit les tuples prêts à l'insertion pour dim_vehicle_avion.
    La valeur CO2 est la moyenne des variants certifiés du même type.
    Les désignations sans valeur CO2 exploitable sont ignorées avec un warning.
    """
    rows = []
    for designation, data in sorted(types.items()):
        try:
            if not data["co2_values"]:
                logger.warning(f"  -> {designation} : aucune valeur CO2 — ignoré.")
                continue

            avg_co2      = round(mean(data["co2_values"]), 4)
            mtom         = data["mtom"]
            service_type = _classify_service(mtom)
            icao_tc      = EASA_TO_ICAO.get(designation)

            rows.append((
                designation,
                avg_co2,
                service_type,
                icao_tc,
                mtom,
                data["engines"],
                data["holder"],
            ))
            logger.info(
                f"  {designation:35s}  CO2={avg_co2:.4f} kg/km  "
                f"service={service_type:22s}  ICAO={icao_tc or '-'}"
            )
        except Exception as e:
            logger.warning(f"  -> {designation} : erreur de transformation ({e}) — ignoré.")

    logger.info(f"[TRANSFORM] EASA : {len(rows)} types d'avions prêts à l'insertion.")
    return rows


def transform_flights(
    table,
    airports_index: dict[str, AirportInfo],
) -> tuple[dict[str, AirportInfo | None], dict[tuple[str, str], RouteInfo]]:
    """
    1. Filtre les vols dont adep ET ades sont présents.
    2. Enrichit chaque aéroport avec ses coordonnées.
    3. Calcule la distance Haversine entre adep et ades.
    4. Identifie le typecode ICAO dominant par route.
    """
    try:
        mask = pc.and_(
            pc.and_(pc.is_valid(table.column("adep")), pc.not_equal(table.column("adep"), "")),
            pc.and_(pc.is_valid(table.column("ades")), pc.not_equal(table.column("ades"), "")),
        )
        filtered = table.filter(mask)
        excluded = len(table) - len(filtered)
        logger.info(
            f"[TRANSFORM] {len(filtered):,} vols valides "
            f"({excluded:,} exclus - adep ou ades manquant)"
        )
    except Exception as e:
        logger.error(f"[TRANSFORM] Erreur lors du filtrage des vols : {e}")
        return {}, {}

    try:
        adep_list = [str(x).strip() for x in filtered.column("adep").to_pylist()]
        ades_list = [str(x).strip() for x in filtered.column("ades").to_pylist()]
    except Exception as e:
        logger.error(f"[TRANSFORM] Erreur lors de la lecture adep/ades : {e}")
        return {}, {}

    # Détection du nom de colonne du type d'avion (varie selon la version OPDI)
    schema_names = table.schema.names
    tc_col = next(
        (c for c in ("aircraft_type", "typecode", "actype", "aircraft_type_icao")
         if c in schema_names),
        None,
    )
    if tc_col:
        logger.info(f"  -> Colonne typecode trouvée : '{tc_col}'")
        try:
            tc_list = [
                str(x).strip()[:20] if x else ""
                for x in filtered.column(tc_col).to_pylist()
            ]
        except Exception as e:
            logger.warning(f"  -> Erreur lecture typecode ({e}) — typecodes ignorés.")
            tc_list = [""] * len(adep_list)
    else:
        logger.warning(
            f"  -> Aucune colonne typecode trouvée. "
            f"Colonnes disponibles : {schema_names}"
        )
        tc_list = [""] * len(adep_list)

    airports:        dict[str, AirportInfo | None]         = {}
    route_distances: dict[tuple[str, str], float | None]   = {}
    route_tc_counts: dict[tuple[str, str], dict[str, int]] = {}

    for adep, ades, tc in zip(adep_list, ades_list, tc_list):
        try:
            if adep not in airports:
                airports[adep] = airports_index.get(adep)
            if ades not in airports:
                airports[ades] = airports_index.get(ades)

            route_key = (adep, ades)

            if route_key not in route_distances:
                info_dep = airports_index.get(adep)
                info_arr = airports_index.get(ades)
                if info_dep and info_arr:
                    route_distances[route_key] = round(
                        _haversine(info_dep[0], info_dep[1], info_arr[0], info_arr[1]), 3
                    )
                else:
                    route_distances[route_key] = None

            if tc:
                if route_key not in route_tc_counts:
                    route_tc_counts[route_key] = {}
                route_tc_counts[route_key][tc] = (
                    route_tc_counts[route_key].get(tc, 0) + 1
                )
        except Exception as e:
            logger.warning(f"  -> Vol {adep}->{ades} ignoré : {e}")

    routes: dict[tuple[str, str], RouteInfo] = {}
    for key in route_distances:
        try:
            routes[key] = (
                route_distances[key],
                max(route_tc_counts[key], key=route_tc_counts[key].get)
                if key in route_tc_counts else None,
            )
        except Exception as e:
            logger.warning(f"  -> Route {key} ignorée lors de la consolidation : {e}")

    with_dist = sum(1 for d, _ in routes.values() if d is not None)
    with_tc   = sum(1 for _, tc in routes.values() if tc is not None)
    logger.info(
        f"           {len(airports):,} aéroports  |  "
        f"{len(routes):,} routes  |  {with_dist:,} avec distance  |  "
        f"{with_tc:,} avec typecode dominant"
    )
    return airports, routes


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 3 — LOAD
# ═══════════════════════════════════════════════════════════════════════════════

def load_vehicle_avion(rows: list[tuple], conn) -> bool:
    """
    a) Ajoute les colonnes manquantes à dim_vehicle_avion (migration idempotente).
    b) Upsert des types d'avions EASA (ON CONFLICT sur label).
    Retourne True si succès, False sinon.
    """
    logger.info(f"[LOAD] dim_vehicle_avion : {len(rows)} type(s) d'avion")
    try:
        with conn.cursor() as cur:
            cur.execute("""
                ALTER TABLE mart.dim_vehicle_avion
                    ADD COLUMN IF NOT EXISTS icao_typecode VARCHAR(10),
                    ADD COLUMN IF NOT EXISTS mtom_kg       INTEGER,
                    ADD COLUMN IF NOT EXISTS num_engines   SMALLINT,
                    ADD COLUMN IF NOT EXISTS manufacturer  VARCHAR(150);
            """)
            cur.execute("""
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM pg_constraint
                        WHERE conname = 'dim_vehicle_avion_label_key'
                    ) THEN
                        ALTER TABLE mart.dim_vehicle_avion
                            ADD CONSTRAINT dim_vehicle_avion_label_key UNIQUE (label);
                    END IF;
                END $$;
            """)
            execute_values(
                cur,
                """
                INSERT INTO mart.dim_vehicle_avion
                    (label, co2_per_km, service_type, icao_typecode,
                     mtom_kg, num_engines, manufacturer)
                VALUES %s
                ON CONFLICT (label) DO UPDATE
                    SET co2_per_km    = EXCLUDED.co2_per_km,
                        service_type  = EXCLUDED.service_type,
                        icao_typecode = EXCLUDED.icao_typecode,
                        mtom_kg       = EXCLUDED.mtom_kg,
                        num_engines   = EXCLUDED.num_engines,
                        manufacturer  = EXCLUDED.manufacturer
                """,
                rows,
            )
            logger.info(f"  -> {len(rows)} type(s) insérés/mis à jour.")
            return True
    except psycopg2.Error as e:
        logger.error(f"[LOAD] dim_vehicle_avion — erreur PostgreSQL : {e}")
    except Exception as e:
        logger.error(f"[LOAD] dim_vehicle_avion — erreur inattendue : {e}")
    return False


def load_stations_and_routes(
    airports: dict[str, AirportInfo | None],
    routes:   dict[tuple[str, str], RouteInfo],
    conn,
) -> bool:
    """
    b) Upsert des aéroports dans mart.dim_station.
    c) Upsert des routes dans mart.dim_route_avion.
       Puis UPDATE co2_total_kg = distance_km x co2_per_km via JOIN EASA.
    Retourne True si succès, False sinon.
    """
    try:
        with conn.cursor() as cur:

            # ── dim_station ──────────────────────────────────────────────────
            station_rows = []
            for code in sorted(airports):
                try:
                    info = airports[code]
                    if info:
                        lat, lon, name, _ = info
                        station_rows.append(
                            (code, name, _get_country(code), True, lat, lon)
                        )
                    else:
                        station_rows.append(
                            (code, code, _get_country(code), True, None, None)
                        )
                except Exception as e:
                    logger.warning(f"  -> Station {code} ignorée : {e}")

            try:
                execute_values(
                    cur,
                    """
                    INSERT INTO mart.dim_station
                        (station_id, station_name, country_code, is_airport,
                         latitude, longitude)
                    VALUES %s
                    ON CONFLICT (station_id) DO UPDATE
                        SET station_name = EXCLUDED.station_name,
                            latitude     = EXCLUDED.latitude,
                            longitude    = EXCLUDED.longitude
                    """,
                    station_rows,
                )
                logger.info(f"[LOAD] dim_station     : {len(station_rows)} aéroport(s)")
            except psycopg2.Error as e:
                logger.error(f"[LOAD] dim_station — erreur PostgreSQL : {e}")
                raise  # remonte pour rollback de la transaction

            # ── dim_route_avion : migration des colonnes ──────────────────────
            try:
                cur.execute("""
                    ALTER TABLE mart.dim_route_avion
                        ADD COLUMN IF NOT EXISTS dominant_typecode VARCHAR(20),
                        ADD COLUMN IF NOT EXISTS co2_total_kg      NUMERIC(12,3);
                """)
                cur.execute("""
                    ALTER TABLE mart.dim_route_avion
                        ALTER COLUMN dominant_typecode TYPE VARCHAR(20);
                """)
            except psycopg2.Error as e:
                logger.error(f"[LOAD] Migration dim_route_avion — erreur : {e}")
                raise

            # ── dim_route_avion : insertion des routes ────────────────────────
            route_rows = [
                (adep, ades, dist, tc)
                for (adep, ades), (dist, tc) in sorted(routes.items())
            ]
            try:
                execute_values(
                    cur,
                    """
                    INSERT INTO mart.dim_route_avion
                        (dep_station_id, arr_station_id, distance_km,
                         dominant_typecode)
                    VALUES %s
                    ON CONFLICT (dep_station_id, arr_station_id) DO UPDATE
                        SET distance_km       = EXCLUDED.distance_km,
                            dominant_typecode = EXCLUDED.dominant_typecode
                    """,
                    route_rows,
                )
                logger.info(f"[LOAD] dim_route_avion : {len(route_rows)} route(s)")
            except psycopg2.Error as e:
                logger.error(f"[LOAD] dim_route_avion — erreur PostgreSQL : {e}")
                raise

            # ── co2_total_kg : JOIN précis (typecode EASA exact) ──────────────
            try:
                cur.execute("""
                    UPDATE mart.dim_route_avion r
                    SET co2_total_kg = ROUND(
                        r.distance_km::NUMERIC * v.co2_per_km, 3
                    )
                    FROM mart.dim_vehicle_avion v
                    WHERE v.icao_typecode = r.dominant_typecode
                      AND r.distance_km   IS NOT NULL;
                """)
                logger.info(
                    f"[LOAD] co2_total_kg    : {cur.rowcount:,} route(s) calculées "
                    f"(typecode EASA exact)"
                )
            except psycopg2.Error as e:
                logger.error(f"[LOAD] co2_total_kg (exact) — erreur : {e}")
                raise

            # ── co2_total_kg : fallback par catégorie de vol (méthode ICAO) ───
            # Pour les routes sans correspondance exacte EASA :
            #   < 500 km   -> regional
            #   < 4000 km  -> court_moyen_courrier
            #   >= 4000 km -> long_courrier
            try:
                cur.execute("""
                    UPDATE mart.dim_route_avion r
                    SET co2_total_kg = ROUND(
                        r.distance_km::NUMERIC * (
                            SELECT AVG(v.co2_per_km)
                            FROM mart.dim_vehicle_avion v
                            WHERE v.service_type = CASE
                                WHEN r.distance_km <  500  THEN 'regional'
                                WHEN r.distance_km < 4000  THEN 'court_moyen_courrier'
                                ELSE                             'long_courrier'
                            END
                        ),
                        3
                    )
                    WHERE r.co2_total_kg IS NULL
                      AND r.distance_km   IS NOT NULL;
                """)
                logger.info(
                    f"[LOAD] co2_total_kg    : {cur.rowcount:,} route(s) calculées "
                    f"(moyenne EASA par catégorie — méthode ICAO)"
                )
            except psycopg2.Error as e:
                logger.error(f"[LOAD] co2_total_kg (fallback) — erreur : {e}")
                raise

        return True

    except psycopg2.Error as e:
        logger.error(f"[LOAD] Stations/Routes — erreur PostgreSQL : {e}")
    except Exception as e:
        logger.error(f"[LOAD] Stations/Routes — erreur inattendue : {e}")
    return False


# ═══════════════════════════════════════════════════════════════════════════════
# ORCHESTRATION
# ═══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    logger.info("=== Pipeline ETL RailCarbon — début ===")

    # ── EXTRACT EASA ──────────────────────────────────────────────────────────
    raw_easa = extract_easa(EASA_FILE, EASA_SHEET)
    if raw_easa is None:
        logger.error("Impossible de lire le fichier EASA — arrêt du pipeline.")
        return

    # ── EXTRACT aéroports ─────────────────────────────────────────────────────
    airports_index = load_airports_index(AIRPORTS_CSV)
    if not airports_index:
        logger.error("Index aéroports vide — arrêt du pipeline.")
        return

    # ── CHECKPOINT : reprise ou démarrage fresh ───────────────────────────────
    checkpoint = load_checkpoint()
    if checkpoint:
        completed_months: list[str]                       = checkpoint["completed_months"]
        all_airports:     dict[str, AirportInfo | None]   = checkpoint["all_airports"]
        all_routes:       dict[tuple[str, str], RouteInfo] = checkpoint["all_routes"]
        all_tc_counts:    dict[tuple[str, str], dict]      = checkpoint["all_tc_counts"]
    else:
        completed_months = []
        all_airports     = {}
        all_routes       = {}
        all_tc_counts    = {}

    # ── EXTRACT + TRANSFORM mensuel ───────────────────────────────────────────
    urls = _generate_urls()
    logger.info(f"{len(urls)} mois à traiter ({START_DATE} -> {END_DATE}).")

    for url in urls:
        month = url.split("_")[-1].replace(".parquet", "")

        if month in completed_months:
            logger.info(f"[SKIP] {month} — déjà traité (checkpoint).")
            continue

        try:
            table = extract_parquet(url)
            if table is None:
                logger.warning(f"[SKIP] {month} — téléchargement échoué, mois ignoré.")
                continue

            airports, routes = transform_flights(table, airports_index)
            del table  # libération mémoire immédiate

            if not airports and not routes:
                logger.warning(f"[SKIP] {month} — aucune donnée extraite.")
                continue

            # Fusion aéroports
            for k, v in airports.items():
                all_airports.setdefault(k, v)

            # Fusion routes : préserve la distance, accumule les typecodes
            for key, (dist, tc) in routes.items():
                if key not in all_routes:
                    all_routes[key] = (dist, tc)
                else:
                    existing_dist, _ = all_routes[key]
                    all_routes[key]  = (existing_dist or dist, tc)

                if tc:
                    if key not in all_tc_counts:
                        all_tc_counts[key] = {}
                    all_tc_counts[key][tc] = all_tc_counts[key].get(tc, 0) + 1

            # Marquer le mois comme traité + sauvegarder le checkpoint
            completed_months.append(month)
            save_checkpoint(completed_months, all_airports, all_routes, all_tc_counts)

        except Exception as e:
            logger.error(f"[ERROR] Mois {month} — erreur inattendue : {e} — mois ignoré.")
            continue

    if not all_routes:
        logger.error("Aucune route extraite — arrêt du pipeline.")
        return

    # Recalcul du typecode dominant sur l'ensemble de l'année
    for key in all_routes:
        try:
            dist, _ = all_routes[key]
            dominant = (
                max(all_tc_counts[key], key=all_tc_counts[key].get)
                if key in all_tc_counts else None
            )
            all_routes[key] = (dist, dominant)
        except Exception as e:
            logger.warning(f"  -> Dominant typecode ignoré pour {key} : {e}")

    with_dist = sum(1 for d, _ in all_routes.values() if d is not None)
    with_tc   = sum(1 for _, t in all_routes.values() if t is not None)
    logger.info(
        f"Cumul annuel : {len(all_airports):,} aéroports | "
        f"{len(all_routes):,} routes | {with_dist:,} avec distance | "
        f"{with_tc:,} avec typecode dominant."
    )

    # ── TRANSFORM EASA ────────────────────────────────────────────────────────
    logger.info("[TRANSFORM] Données EASA...")
    easa_rows = transform_easa(raw_easa)
    if not easa_rows:
        logger.error("Aucune donnée EASA transformée — arrêt du pipeline.")
        return

    # ── LOAD ──────────────────────────────────────────────────────────────────
    logger.info(
        f"[LOAD] Connexion PostgreSQL ({DB_CONFIG['host']}:{DB_CONFIG['port']})..."
    )
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except psycopg2.OperationalError as e:
        logger.error(f"[LOAD] Connexion impossible : {e}")
        logger.error("Vérifiez que le conteneur est démarré : docker compose up -d")
        return

    load_ok = False
    try:
        with conn:  # transaction atomique : commit si tout réussit, rollback sinon
            easa_ok  = load_vehicle_avion(easa_rows, conn)
            if not easa_ok:
                raise RuntimeError("Echec chargement dim_vehicle_avion — rollback.")

            routes_ok = load_stations_and_routes(all_airports, all_routes, conn)
            if not routes_ok:
                raise RuntimeError("Echec chargement stations/routes — rollback.")

            load_ok = True

    except (psycopg2.Error, RuntimeError) as e:
        logger.error(f"[LOAD] Transaction annulée (rollback) : {e}")
    except Exception as e:
        logger.error(f"[LOAD] Erreur inattendue — rollback : {e}")
    finally:
        conn.close()
        logger.info("[LOAD] Connexion fermée.")

    # ── FIN ───────────────────────────────────────────────────────────────────
    if load_ok:
        clear_checkpoint()  # supprime le checkpoint uniquement si tout a réussi
        logger.info("=== Pipeline ETL RailCarbon — terminé avec succès ===")
    else:
        logger.error(
            "=== Pipeline ETL RailCarbon — terminé avec erreurs "
            "(checkpoint conservé pour reprise) ==="
        )


if __name__ == "__main__":
    main()
