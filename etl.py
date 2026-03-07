"""
etl.py
======
Pipeline ETL UNIQUE — Train (GTFS + Back on Track) + Avion (OPDI + EASA).
Entièrement sur PySpark (une seule session partagée entre les deux sections).

Architecture : chaque mode de transport suit strictement les 3 phases ETL,
               dans des sections clairement délimitées.

  ╔══════════════════════════════════════════════════════════════════╗
  ║  SECTION TRAIN                                                   ║
  ║  ├── EXTRACT   — Téléchargement GTFS + Back on Track            ║
  ║  ├── TRANSFORM — Déduplication stations, routes, détection TGV  ║
  ║  └── LOAD      — dim_station, dim_route_train, fact_emission     ║
  ╠══════════════════════════════════════════════════════════════════╣
  ║  SECTION AVION                                                   ║
  ║  ├── EXTRACT   — EASA Excel + airports.csv + OPDI parquets      ║
  ║  ├── TRANSFORM — Haversine, typecode dominant, agrégation EASA  ║
  ║  └── LOAD      — dim_vehicle_avion, dim_station, dim_route_avion ║
  ╚══════════════════════════════════════════════════════════════════╝

Infrastructure transverse :
  - PySpark     : session unique partagée train + avion
  - Loki        : logs Python → Grafana en temps réel (optionnel)
  - etl_run_log : traçabilité de chaque run dans PostgreSQL (via psql)
  - opdi_cache/ : cache disque des parquets mensuels (remplace pickle checkpoint)

Workflow :
  docker compose up -d
  python etl.py
"""

# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════

# ── stdlib ────────────────────────────────────────────────────────────────────
import gc
import os
import re
import psycopg2
import urllib.request
import sys
import tempfile
import zipfile
import logging
from collections import defaultdict
from datetime import datetime
from math import asin as _m_asin
from math import cos as _m_cos
from math import radians as _m_radians
from math import sin as _m_sin
from math import sqrt as _m_sqrt
from statistics import mean

from dateutil.relativedelta import relativedelta

# ── tiers — EASA (openpyxl) + HTTP (requests) ────────────────────────────────
# openpyxl : seul outil tiers non-PySpark conservé (PySpark ne lit pas l'Excel)
import openpyxl
import requests

# ── PySpark — session unique train + avion ────────────────────────────────────
try:
    from pyspark.sql import SparkSession
    from pyspark.sql.functions import (
        asin as sp_asin,
        avg as sp_avg,
        broadcast,
        coalesce,
        col,
        collect_list,
        concat,
        concat_ws,
        cos as sp_cos,
        count as sp_count,
        lead,
        lit,
        md5,
        radians as sp_radians,
        row_number,
        sin as sp_sin,
        sort_array,
        sqrt as sp_sqrt,
        upper,
        when,
    )
    from pyspark.sql.window import Window
    from pyspark.sql.types import FloatType, IntegerType, StringType
except ImportError:
    print(
        "ERREUR : PySpark non installé.\n"
        "Installation : pip install pyspark\n"
        "Java 11 ou 17 requis (pas Java 25)."
    )
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION COMMUNE
# ═══════════════════════════════════════════════════════════════════════════════

SCHEMA = "mart"

LOG_FILE = f"etl_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
import io as _io
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(
            _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        ),
    ],
)
logger = logging.getLogger(__name__)

# ── Handler Loki (optionnel) ──────────────────────────────────────────────────
# Installation : pip install python-logging-loki
LOKI_URL = os.environ.get("LOKI_URL", "http://localhost:3100/loki/api/v1/push")
try:
    import logging_loki
    _loki_handler = logging_loki.LokiHandler(
        url=LOKI_URL,
        tags={"application": "railcarbon-etl"},
        version="1",
    )
    _loki_handler.setLevel(logging.INFO)
    logger.addHandler(_loki_handler)
    logger.info("[LOGGING] Handler Loki activé -> %s", LOKI_URL)
except ImportError:
    logger.info("[LOGGING] python-logging-loki non installé — logs Loki désactivés.")
except Exception as _e:
    logger.warning("[LOGGING] Loki non disponible (%s) — logs fichier/console uniquement.", _e)

# ── Connexion PostgreSQL ───────────────────────────────────────────────────────
DB_CONFIG = {
    "host":     os.environ.get("DB_HOST",     "localhost"),
    "port":     os.environ.get("DB_PORT",     "5433"),
    "database": os.environ.get("DB_NAME",     "postgres"),
    "user":     os.environ.get("DB_USER",     "postgres"),
    "password": os.environ.get("DB_PASSWORD", "lpironti"),
}




# ═══════════════════════════════════════════════════════════════════════════════
# HELPER PSQL  (partagé train + avion)
# ═══════════════════════════════════════════════════════════════════════════════

def _get_pg_conn():
    """Retourne une connexion psycopg2 au PostgreSQL."""
    return psycopg2.connect(
        host=DB_CONFIG["host"],
        port=int(DB_CONFIG["port"]),
        dbname=DB_CONFIG["database"],
        user=DB_CONFIG["user"],
        password=DB_CONFIG["password"],
    )


def _run_sql(sql_command: str) -> bool:
    """Exécute une commande SQL via psycopg2. Retourne True si succès."""
    try:
        conn = _get_pg_conn()
        conn.autocommit = True
        with conn.cursor() as cur:
            cur.execute(sql_command)
        conn.close()
        return True
    except Exception as e:
        logger.error("[SQL] Erreur psycopg2 : %s", str(e)[:300])
        return False


def _run_sql_file(tmp_path: str) -> bool:
    """Exécute un fichier SQL (multi-instructions) via psycopg2. Retourne True si succès."""
    try:
        with open(tmp_path, "r", encoding="utf-8") as f:
            sql = f.read()
        conn = _get_pg_conn()
        conn.autocommit = True
        with conn.cursor() as cur:
            for stmt in sql.split(";\n"):
                stmt = stmt.strip()
                if stmt:
                    cur.execute(stmt)
        conn.close()
        return True
    except Exception as e:
        logger.error("[SQL] Erreur psycopg2 (fichier) : %s", str(e)[:300])
        return False


def _query_sql(sql_command: str) -> str:
    """Exécute une requête et retourne la première valeur scalaire (ex: RETURNING id)."""
    try:
        conn = _get_pg_conn()
        conn.autocommit = True
        with conn.cursor() as cur:
            cur.execute(sql_command)
            row = cur.fetchone()
        conn.close()
        if row:
            return str(row[0])
        return ""
    except Exception as e:
        logger.error("[SQL] Erreur query psycopg2 : %s", str(e)[:200])
        return ""


# ═══════════════════════════════════════════════════════════════════════════════
# PROVENANCE  (mart.dim_source — via psql)
# ═══════════════════════════════════════════════════════════════════════════════

# Sources statiques connues à l'avance (avion + back on track)
_STATIC_SOURCES = [
    {
        "source_name": "EASA CO2 Database",
        "source_type": "fichier_local",
        "url":         "https://www.easa.europa.eu/en/domains/environment/easa-co2-database",
        "version":     "Issue 11",
        "description": "Base CO2 EASA — types d'avions certifiés et émissions kg/km",
    },
    {
        "source_name": "OPDI Eurocontrol",
        "source_type": "api",
        "url":         "https://www.eurocontrol.int/performance/data/download/OPDI/v002/flight_list/",
        "version":     "v002",
        "description": "Données de vols Eurocontrol 2025 — parquets mensuels flight_list",
    },
    {
        "source_name": "OurAirports",
        "source_type": "fichier_local",
        "url":         "https://ourairports.com/data/airports.csv",
        "version":     None,
        "description": "Index mondial des aéroports avec coordonnées GPS (OurAirports open data)",
    },
    {
        "source_name": "Back on Track",
        "source_type": "fichier_local",
        "url":         "https://github.com/back-on-track-europe/open-data",
        "version":     None,
        "description": "Données trains de nuit européens — open data Back on Track Europe",
    },
]


def _source_upsert(
    source_name: str,
    source_type: str,
    url: str = None,
    version: str = None,
    description: str = None,
) -> "int | None":
    """
    Upsert une entrée dans mart.dim_source.
    Crée la table si elle n'existe pas (migration idempotente).
    Retourne le source_id (int) ou None en cas d'erreur.
    """
    _run_sql("""
        CREATE TABLE IF NOT EXISTS mart.dim_source (
            source_id    SERIAL PRIMARY KEY,
            source_name  VARCHAR(200) NOT NULL UNIQUE,
            source_type  VARCHAR(50)  NOT NULL,
            url          TEXT,
            version      VARCHAR(100),
            description  TEXT,
            loaded_at    TIMESTAMP NOT NULL DEFAULT NOW()
        );
    """)

    def esc(v):
        return "NULL" if v is None else f"'{str(v).replace(chr(39), chr(39)*2)}'"

    sql = (
        f"INSERT INTO mart.dim_source (source_name, source_type, url, version, description) "
        f"VALUES ({esc(source_name)}, {esc(source_type)}, {esc(url)}, {esc(version)}, {esc(description)}) "
        f"ON CONFLICT (source_name) DO UPDATE "
        f"SET source_type=EXCLUDED.source_type, url=EXCLUDED.url, "
        f"version=EXCLUDED.version, description=EXCLUDED.description, loaded_at=NOW() "
        f"RETURNING source_id;"
    )
    raw = _query_sql(sql)
    if raw:
        try:
            return int(raw.strip())
        except ValueError:
            pass
    logger.warning("[SOURCE] Impossible d'upsert la source '%s'.", source_name)
    return None


def source_seed_static() -> dict:
    """
    Upsert les sources statiques (EASA, OPDI, OurAirports, Back on Track).
    Retourne un dict {source_name: source_id}.
    """
    ids = {}
    for s in _STATIC_SOURCES:
        sid = _source_upsert(**s)
        if sid:
            ids[s["source_name"]] = sid
            logger.info("[SOURCE] %-30s -> source_id=%d", s["source_name"], sid)
    return ids


def source_upsert_gtfs_feed(feed_id: str, country_code: str, url: str = None) -> "int | None":
    """Upsert une source GTFS par feed. Retourne le source_id."""
    return _source_upsert(
        source_name=f"GTFS - {country_code} - {feed_id}",
        source_type="gtfs_feed",
        url=url,
        version=None,
        description=f"Feed GTFS {country_code}/{feed_id} (MobilityDatabase)",
    )


# ═══════════════════════════════════════════════════════════════════════════════
# ETL RUN LOG  (mart.etl_run_log — via psql)
# ═══════════════════════════════════════════════════════════════════════════════

def run_start() -> "str | None":
    """
    Crée mart.etl_run_log si absente (migration idempotente),
    insère un run 'en_cours'. Retourne le run_id (str) ou None en cas d'échec.
    """
    _run_sql("""
        CREATE TABLE IF NOT EXISTS mart.etl_run_log (
            run_id           SERIAL PRIMARY KEY,
            started_at       TIMESTAMP   NOT NULL DEFAULT NOW(),
            finished_at      TIMESTAMP,
            duration_s       NUMERIC(10,2),
            mois_traites     INTEGER     DEFAULT 0,
            airports_charges INTEGER     DEFAULT 0,
            routes_chargees  INTEGER     DEFAULT 0,
            co2_exact        INTEGER     DEFAULT 0,
            co2_fallback     INTEGER     DEFAULT 0,
            erreurs          INTEGER     DEFAULT 0,
            statut           VARCHAR(20) DEFAULT 'en_cours'
        );
    """)
    run_id = _query_sql(
        "INSERT INTO mart.etl_run_log (started_at, statut) "
        "VALUES (NOW(), 'en_cours') RETURNING run_id;"
    )
    if run_id:
        logger.info("[RUN] ETL run #%s démarré — suivi dans mart.etl_run_log.", run_id)
        return run_id
    logger.warning("[RUN] Impossible de créer le run log — pipeline continue.")
    return None


def run_end(run_id: "str | None", metrics: dict, statut: str) -> None:
    """Met à jour la ligne du run avec les métriques finales (train + avion cumulés)."""
    if run_id is None:
        return
    _run_sql(f"""
        UPDATE mart.etl_run_log
        SET finished_at      = NOW(),
            duration_s       = ROUND(EXTRACT(EPOCH FROM (NOW() - started_at))::NUMERIC, 2),
            mois_traites     = {metrics.get('mois_traites', 0)},
            airports_charges = {metrics.get('airports_charges', 0)},
            routes_chargees  = {metrics.get('routes_chargees', 0)},
            co2_exact        = {metrics.get('co2_exact', 0)},
            co2_fallback     = {metrics.get('co2_fallback', 0)},
            erreurs          = {metrics.get('erreurs', 0)},
            statut           = '{statut}'
        WHERE run_id = {run_id};
    """)
    logger.info(
        "[RUN] ETL run #%s terminé — statut : %s | routes=%d | mois=%d | erreurs=%d",
        run_id, statut,
        metrics.get("routes_chargees", 0),
        metrics.get("mois_traites", 0),
        metrics.get("erreurs", 0),
    )


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║                          SECTION  TRAIN                                  ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

# ── TRAIN — CONFIGURATION ─────────────────────────────────────────────────────

GTFS_DIR        = "downloads_gtfs"
FEEDS_CSV       = "feeds_v2.csv"
BACKONTRACK_DIR = "backontrack"

# ── Pays européens (ISO 3166-1 alpha-2) ──────────────────────
EUROPEAN_COUNTRIES = {
     'AL', 'AD', 'AT', 'BY', 'BE', 'BA', 'BG', 'HR', 'CY', 'CZ',
     'DK', 'EE', 'FI', 'FR', 'DE', 'GR', 'HU', 'IS', 'IE', 'IT',
     'XK', 'LV', 'LI', 'LT', 'LU', 'MT', 'MD', 'MC', 'ME', 'NL',
     'MK', 'NO', 'PL', 'PT', 'RO', 'RU', 'SM', 'RS', 'SK', 'SI',
     'ES', 'SE', 'CH', 'UA', 'GB', 'VA'
 }


TGV_ROUTE_TYPES = {101}

TGV_KEYWORDS = [
    "TGV", "OUIGO", "INOUI",
    "ICE",
    "EUROSTAR", "THALYS",
    "AVE",
    "FRECCIAROSSA", "FRECCIARGENTO",
    "ITALO",
    "RAILJET",
    "FYRA",
    "PENDOLINO",
    "SJ 2000",
]

INTERCITE_ROUTE_TYPES   = {2, 100, 102, 103, 105, 106}
ALL_TRAIN_ROUTE_TYPES   = TGV_ROUTE_TYPES | INTERCITE_ROUTE_TYPES

CO2_TGV_PER_100KM       = 0.29    # kg / 100 km (ADEME — TGV électrique)
CO2_INTERCITE_PER_100KM = 0.90    # kg / 100 km (ADEME — Intercité diesel)
CO2_NIGHTTRAIN_PER_100KM = 1.25   # kg / 100 km (trains de nuit — estimé Back on Track / EEA)

STATION_DEDUP_KM        = 0.15    # 150 m → même gare physique
MAX_NAME_DEDUP_KM       = 50.0    # 50 km max pour fusionner des homonymes

MIN_LAT, MAX_LAT        = -90.0, 90.0
MIN_LON, MAX_LON        = -180.0, 180.0
COORD_ZERO_THRESHOLD    = 0.1

TABLES_TRAIN_TRUNCATE = [
    f"{SCHEMA}.dim_route_train",
    f"{SCHEMA}.dim_vehicle_train",
    f"{SCHEMA}.dim_station",
]

TABLES_AVION_TRUNCATE = [
    f"{SCHEMA}.dim_route_avion",
    f"{SCHEMA}.dim_vehicle_avion",
]


# ── TRAIN — HELPERS SPARK ────────────────────────────────────────────────────

def _haversine_col(lat1, lon1, lat2, lon2):
    """Distance haversine en km — colonnes Spark (partagée train + avion)."""
    R    = 6371.0
    dlat = sp_radians(lat2 - lat1)
    dlon = sp_radians(lon2 - lon1)
    a    = sp_sin(dlat / 2) ** 2 + sp_cos(sp_radians(lat1)) * sp_cos(sp_radians(lat2)) * sp_sin(dlon / 2) ** 2
    return lit(2 * R) * sp_asin(sp_sqrt(a))


def _haversine_py(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Distance haversine en km — Python pur (Union-Find train)."""
    R    = 6371.0
    dlat = _m_radians(lat2 - lat1)
    dlon = _m_radians(lon2 - lon1)
    a    = _m_sin(dlat / 2) ** 2 + _m_cos(_m_radians(lat1)) * _m_cos(_m_radians(lat2)) * _m_sin(dlon / 2) ** 2
    return 2 * R * _m_asin(_m_sqrt(a))


def _is_valid_coord(lat_col, lon_col):
    """Condition Spark filtrant les coordonnées GPS invalides."""
    return (
        lat_col.isNotNull() & lon_col.isNotNull()
        & ~(
            lat_col.between(-COORD_ZERO_THRESHOLD, COORD_ZERO_THRESHOLD)
            & lon_col.between(-COORD_ZERO_THRESHOLD, COORD_ZERO_THRESHOLD)
        )
        & lat_col.between(MIN_LAT, MAX_LAT)
        & lon_col.between(MIN_LON, MAX_LON)
    )


def _read_gtfs_file(spark, feed_path: str, filename: str):
    filepath = os.path.join(feed_path, filename)
    if os.path.exists(filepath):
        # inferSchema=False -> tout en StringType, évite que stop_id numérique soit inféré BIGINT
        df = spark.read.csv(filepath, header=True, inferSchema=False)
        # Certains feeds GTFS ont des espaces en tête dans les noms de colonnes (" route_type")
        return df.toDF(*[c.strip() for c in df.columns])
    return None


# ── TRAIN — EXTRACT ───────────────────────────────────────────────────────────

def train_download_all_gtfs(spark) -> None:
    """Télécharge tous les feeds GTFS européens depuis feeds_v2.csv."""
    os.makedirs(GTFS_DIR, exist_ok=True)
    df = spark.read.csv(FEEDS_CSV, header=True, inferSchema=True)
    df_europe = df.filter(
        (col("`location.country_code`").isin(list(EUROPEAN_COUNTRIES)))
        & col("`urls.latest`").isNotNull()
        & (col("status") == "active")
    ).select(
        col("id"),
        col("`location.country_code`").alias("country_code"),
        col("`urls.latest`").alias("url"),
    )
    logger.info("[TRAIN][EXTRACT] %d feeds européens trouvés.", df_europe.count())
    feeds = df_europe.collect()

    def _download_one_feed(feed):
        feed_dir = os.path.join(GTFS_DIR, feed["country_code"], str(feed["id"]))
        if os.path.isdir(feed_dir):
            return True
        try:
            response = requests.get(feed["url"], timeout=30, stream=True)
            response.raise_for_status()
            country_dir = os.path.join(GTFS_DIR, feed["country_code"])
            os.makedirs(country_dir, exist_ok=True)
            tmp_zip = os.path.join(country_dir, f"{feed['id']}.zip")
            with open(tmp_zip, "wb") as f:
                for chunk in response.iter_content(chunk_size=65536):
                    f.write(chunk)
            extract_dir = os.path.join(country_dir, str(feed["id"]))
            os.makedirs(extract_dir, exist_ok=True)
            with zipfile.ZipFile(tmp_zip, "r") as z:
                z.extractall(extract_dir)
            os.remove(tmp_zip)
            logger.info("[TRAIN][EXTRACT] Telecharge : %s/%s", feed["country_code"], feed["id"])
            return True
        except Exception as e:
            logger.error("[TRAIN][EXTRACT] Erreur %s/%s : %s", feed["country_code"], feed["id"], e)
            return False

    from concurrent.futures import ThreadPoolExecutor, as_completed
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(_download_one_feed, f): f for f in feeds}
        success = sum(1 for fut in as_completed(futures) if fut.result())

    logger.info("[TRAIN][EXTRACT] %d/%d feeds telecharges.", success, len(feeds))


def train_list_gtfs_feeds(base_dir: str) -> list:
    """Liste tous les dossiers GTFS disponibles."""
    feeds = []
    for country_code in os.listdir(base_dir):
        country_path = os.path.join(base_dir, country_code)
        if os.path.isdir(country_path):
            for feed_id in os.listdir(country_path):
                feed_path = os.path.join(country_path, feed_id)
                if os.path.isdir(feed_path):
                    feeds.append({"country_code": country_code, "feed_id": feed_id, "path": feed_path})
    return feeds


def train_get_stop_ids(spark, feed_path: str) -> set:
    """Retourne l'ensemble des stop_id utilisés par des trips ferroviaires dans ce feed."""
    routes_df     = _read_gtfs_file(spark, feed_path, "routes.txt")
    trips_df      = _read_gtfs_file(spark, feed_path, "trips.txt")
    stop_times_df = _read_gtfs_file(spark, feed_path, "stop_times.txt")
    if routes_df is None or trips_df is None or stop_times_df is None:
        return set()
    train_routes = routes_df.filter(
        col("route_type").cast(IntegerType()).isin(list(ALL_TRAIN_ROUTE_TYPES))
    ).select("route_id")
    if not train_routes.head(1):
        return set()
    train_trips    = trips_df.join(train_routes, "route_id", "semi").select("trip_id")
    train_stop_ids = (
        stop_times_df.join(train_trips, "trip_id", "semi")
        .select(col("stop_id").cast(StringType())).distinct()
    )
    return set(r.stop_id for r in train_stop_ids.collect())


def train_collect_stations_from_feed(spark, feed_path: str, country_code: str, train_stop_ids: set, source_id: "int | None" = None) -> list:
    """Extrait les métadonnées de stations d'un feed GTFS (résolution parent_station)."""
    df = _read_gtfs_file(spark, feed_path, "stops.txt")
    if df is None:
        return []
    if "location_type" in df.columns:
        df = df.filter((col("location_type").isNull()) | (col("location_type").isin([0, 1])))
    df = df.withColumn("lat", col("stop_lat").cast(FloatType())) \
           .withColumn("lon", col("stop_lon").cast(FloatType()))
    df = df.filter(_is_valid_coord(col("lat"), col("lon")))
    has_parent  = "parent_station" in df.columns
    select_cols = [
        col("stop_id").cast(StringType()).alias("stop_id"),
        col("stop_name").cast(StringType()).alias("station_name"),
        col("lat").alias("latitude"),
        col("lon").alias("longitude"),
    ]
    if has_parent:
        select_cols.append(col("parent_station").cast(StringType()).alias("parent_station"))
    rows       = df.select(*select_cols).collect()
    stop_lookup = {row["stop_id"]: row for row in rows}
    expanded   = set(train_stop_ids)
    if has_parent:
        for sid in train_stop_ids:
            if sid in stop_lookup:
                parent = stop_lookup[sid]["parent_station"]
                if parent and parent in stop_lookup:
                    expanded.add(parent)
    result = []
    seen   = set()
    for sid in expanded:
        if sid not in stop_lookup:
            continue
        row          = stop_lookup[sid]
        resolved_id  = sid
        resolved_row = row
        if has_parent and row["parent_station"]:
            pid = row["parent_station"]
            if pid in stop_lookup:
                resolved_id  = pid
                resolved_row = stop_lookup[pid]
        if resolved_id in seen:
            continue
        seen.add(resolved_id)
        name = resolved_row["station_name"] or ""
        result.append({
            "resolved_stop_id": resolved_id,
            "station_name":     name,
            "city":             _extract_city(name),
            "normalized_name":  _train_normalize_name(name),
            "latitude":         float(resolved_row["latitude"]),
            "longitude":        float(resolved_row["longitude"]),
            "country_code":     country_code,
            "source_id":        source_id,
        })
    return result


def train_collect_stations_from_backontrack(spark, source_id: "int | None" = None) -> list:
    """Collecte les stations Back on Track (trains de nuit européens)."""
    stops_path     = os.path.join(BACKONTRACK_DIR, "back_on_track_stops.csv")
    trip_stop_path = os.path.join(BACKONTRACK_DIR, "back_on_track_trip_stop.csv")
    if not os.path.exists(stops_path) or not os.path.exists(trip_stop_path):
        logger.warning("[TRAIN][EXTRACT] Back on Track : fichiers manquants, ignoré.")
        return []
    trip_stop_df  = spark.read.csv(trip_stop_path, header=True, inferSchema=True)
    used_stop_ids = set(
        r.stop_id
        for r in trip_stop_df.select(col("stop_id").cast(StringType())).distinct().collect()
    )
    stops_df = (
        spark.read.csv(stops_path, header=True, inferSchema=True)
        .withColumn("lat", col("station_lat").cast(FloatType()))
        .withColumn("lon", col("station_long").cast(FloatType()))
    )
    stops_df = stops_df.filter(_is_valid_coord(col("lat"), col("lon")))
    rows     = stops_df.select(
        col("stop_id").cast(StringType()).alias("stop_id"),
        col("station_name").cast(StringType()).alias("station_name"),
        col("lat").alias("latitude"),
        col("lon").alias("longitude"),
        col("stop_country").cast(StringType()).alias("country_code"),
    ).collect()
    result = []
    for row in rows:
        sid  = row["stop_id"]
        if sid not in used_stop_ids:
            continue
        name = row["station_name"] or sid
        result.append({
            "resolved_stop_id": sid,
            "station_name":     name,
            "normalized_name":  _train_normalize_name(name),
            "latitude":         float(row["latitude"]),
            "longitude":        float(row["longitude"]),
            "country_code":     row["country_code"] or "",
            "source_id":        source_id,
        })
    return result


# ── TRAIN — TRANSFORM ─────────────────────────────────────────────────────────

def _extract_city(station_name: str) -> str:
    """Extrait la ville depuis le nom d'une gare (heuristique).
    Ex: 'Paris - Gare de Lyon' → 'Paris', 'Lyon, Part-Dieu' → 'Lyon'."""
    if not station_name:
        return ""
    for sep in [" - ", ", ", " ("]:
        if sep in station_name:
            return station_name.split(sep)[0].strip()
    return station_name.strip()


def _train_normalize_name(name: str) -> str:
    if not name:
        return ""
    name = re.sub(r"\s*[\(\[][A-Z0-9]{2,5}[\)\]]\s*$", "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def _train_deduplicate_stations(stations_rows: list) -> dict:
    """Union-Find géographique : fusionne les stations à moins de STATION_DEDUP_KM."""
    stations = [
        {"station_id": r["station_id"], "station_name": r["station_name"] or "",
         "latitude": float(r["latitude"]), "longitude": float(r["longitude"])}
        for r in stations_rows
    ]
    n = len(stations)
    if n == 0:
        return {}
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(x, y):
        px, py = find(x), find(y)
        if px != py:
            parent[px] = py

    sorted_idx = sorted(range(n), key=lambda i: stations[i]["latitude"])
    for ii in range(len(sorted_idx)):
        i     = sorted_idx[ii]
        lat_i = stations[i]["latitude"]
        lon_i = stations[i]["longitude"]
        for jj in range(ii + 1, len(sorted_idx)):
            j     = sorted_idx[jj]
            lat_j = stations[j]["latitude"]
            if abs(lat_j - lat_i) > 0.01:
                break
            if _haversine_py(lat_i, lon_i, lat_j, stations[j]["longitude"]) < STATION_DEDUP_KM:
                union(i, j)

    clusters = {}
    for i in range(n):
        clusters.setdefault(find(i), []).append(i)

    mapping = {}
    for members in clusters.values():
        if len(members) == 1:
            continue
        members.sort(key=lambda i: (-len(stations[i]["station_name"]), stations[i]["station_id"]))
        canonical = stations[members[0]]["station_id"]
        for m in members[1:]:
            mapping[stations[m]["station_id"]] = canonical

    if mapping:
        logger.info("[TRAIN][TRANSFORM] Dédup géo : %d doublons fusionnés.", len(mapping))
    return mapping


def train_build_station_registry(all_stations: list) -> tuple:
    """
    Déduplication en 2 couches : nom normalisé (< MAX_NAME_DEDUP_KM) puis
    géographique Union-Find (< STATION_DEDUP_KM).
    Retourne (canonical_stations: list, global_mapping: dict {old_id: canonical_id}).
    """
    name_groups = defaultdict(list)
    for s in all_stations:
        key = s["normalized_name"].lower()
        name_groups[key if key else f"__empty_{s['resolved_stop_id']}"].append(s)

    merged = []
    name_dedup_count = 0
    for group in name_groups.values():
        if len(group) == 1:
            merged.append(group[0])
            continue
        sub_groups = []
        used = [False] * len(group)
        for i in range(len(group)):
            if used[i]:
                continue
            cluster = [i]
            used[i] = True
            for j in range(i + 1, len(group)):
                if used[j]:
                    continue
                if _haversine_py(
                    group[i]["latitude"], group[i]["longitude"],
                    group[j]["latitude"], group[j]["longitude"],
                ) < MAX_NAME_DEDUP_KM:
                    cluster.append(j)
                    used[j] = True
            sub_groups.append(cluster)
        for cl in sub_groups:
            members = [group[i] for i in cl]
            members.sort(key=lambda s: (-len(s["station_name"]), s["resolved_stop_id"]))
            canonical = members[0].copy()
            canonical["merged_ids"] = {s["resolved_stop_id"] for s in members}
            merged.append(canonical)
            if len(members) > 1:
                name_dedup_count += len(members) - 1

    if name_dedup_count:
        logger.info("[TRAIN][TRANSFORM] Dédup par nom : %d doublons fusionnés.", name_dedup_count)

    geo_mapping = _train_deduplicate_stations([
        {"station_id": s["resolved_stop_id"], "station_name": s["station_name"],
         "latitude": s["latitude"], "longitude": s["longitude"]}
        for s in merged
    ])

    canonical_by_id = {}
    for s in merged:
        rid      = s["resolved_stop_id"]
        final_id = geo_mapping.get(rid, rid)
        if final_id not in canonical_by_id:
            canonical_by_id[final_id] = s.copy()
            canonical_by_id[final_id]["resolved_stop_id"] = final_id

    global_mapping = {}
    for s in merged:
        rid      = s["resolved_stop_id"]
        final_id = geo_mapping.get(rid, rid)
        if rid != final_id:
            global_mapping[rid] = final_id
        for mid in s.get("merged_ids", set()):
            if mid != final_id:
                global_mapping[mid] = final_id

    canonical_stations = [
        {"station_id": cid, "station_name": s["station_name"],
         "city": s.get("city") or _extract_city(s["station_name"]),
         "country_code": s["country_code"], "latitude": s["latitude"],
         "longitude": s["longitude"], "is_airport": False,
         "source_id": s.get("source_id")}
        for cid, s in canonical_by_id.items()
    ]
    logger.info(
        "[TRAIN][TRANSFORM] Registre : %d stations canoniques, %d remappings.",
        len(canonical_stations), len(global_mapping),
    )
    return canonical_stations, global_mapping


def train_build_feed_stop_mapping(spark, feed_path: str, global_mapping: dict) -> dict:
    """Mapping complet stop_id → canonical_station_id pour un feed."""
    df = _read_gtfs_file(spark, feed_path, "stops.txt")
    if df is None:
        return {}
    has_parent  = "parent_station" in df.columns
    select_cols = [col("stop_id").cast(StringType()).alias("stop_id")]
    if has_parent:
        select_cols.append(col("parent_station").cast(StringType()).alias("parent_station"))
    mapping = {}
    for row in df.select(*select_cols).collect():
        sid      = row["stop_id"]
        resolved = sid
        if has_parent and row["parent_station"]:
            resolved = row["parent_station"]
        final = global_mapping.get(resolved, resolved)
        if sid in global_mapping:
            final = global_mapping[sid]
        if sid != final:
            mapping[sid] = final
    return mapping


def _build_stop_pairs_with_distance(stop_times_df, stops_clean, extra_cols=()):
    """
    Helper partagé train_process_routes_and_facts + train_process_backontrack.
    Déduplique les trips par signature MD5, extrait les tronçons consécutifs dep→arr,
    joint avec les coordonnées GPS (stops_clean) et calcule la distance haversine.

    stop_times_df : DataFrame(trip_id, stop_id, stop_sequence[, ...extra_cols])
    stops_clean   : DataFrame(stop_id, lat, lon) — déjà castées et filtrées
    extra_cols    : colonnes additionnelles à conserver dans le résultat (ex: ("route_id",))

    Retourne DataFrame(dep_stop_id, arr_stop_id, distance_km[, ...extra_cols]).
    """
    group_cols = ["trip_id"] + list(extra_cols)
    trip_signatures = (
        stop_times_df
        .groupBy(*group_cols)
        .agg(sort_array(collect_list(
            concat(col("stop_sequence").cast(StringType()), lit(":"),
                   col("stop_id").cast(StringType()))
        )).alias("_stops_ordered"))
        .withColumn("trip_signature", md5(concat_ws("|", col("_stops_ordered"))))
        .drop("_stops_ordered")
        .dropDuplicates(["trip_signature"])
    )
    trip_signatures.cache()

    w_trip = Window.partitionBy("trip_id").orderBy(col("stop_sequence").cast(FloatType()))
    stop_pairs = (
        stop_times_df
        .join(trip_signatures.select("trip_id"), "trip_id", "semi")
        .withColumn("_next_stop_id", lead("stop_id").over(w_trip))
        .filter(col("_next_stop_id").isNotNull())
        .select(
            col("stop_id").cast(StringType()).alias("dep_stop_id"),
            col("_next_stop_id").cast(StringType()).alias("arr_stop_id"),
            *[col(c) for c in extra_cols],
        )
        .filter(col("dep_stop_id") != col("arr_stop_id"))
    )

    pairs_with_dist = (
        stop_pairs
        .join(
            stops_clean.withColumnRenamed("stop_id", "dep_stop_id")
                       .withColumnRenamed("lat", "dep_lat")
                       .withColumnRenamed("lon", "dep_lon"),
            "dep_stop_id", "inner",
        )
        .join(
            stops_clean.withColumnRenamed("stop_id", "arr_stop_id")
                       .withColumnRenamed("lat", "arr_lat")
                       .withColumnRenamed("lon", "arr_lon"),
            "arr_stop_id", "inner",
        )
        .withColumn("distance_km",
                    _haversine_col(col("dep_lat"), col("dep_lon"),
                                   col("arr_lat"), col("arr_lon")))
    )
    trip_signatures.unpersist()
    return pairs_with_dist


def train_process_routes_and_facts(
    spark, feed_path: str, feed_id: str, country_code: str,
    station_id_mapping: dict = None,
) -> tuple:
    """
    Pour chaque trip ferroviaire : tronçons consécutifs → distance haversine → CO2.
    Retourne (unique_routes_df, facts_with_co2_df) ou (None, None).
    """
    if station_id_mapping is None:
        station_id_mapping = {}

    routes_df = _read_gtfs_file(spark, feed_path, "routes.txt")
    if routes_df is None:
        return None, None
    routes_df = routes_df.filter(
        col("route_type").rlike(r"^\s*\d+\s*$")  # exclut 'TransporteAereo' et autres non-numériques
    ).filter(
        col("route_type").cast(IntegerType()).isin(list(ALL_TRAIN_ROUTE_TYPES))
    )
    if not routes_df.head(1):
        return None, None

    route_name_cols = ["route_id", "route_type"]
    if "route_short_name" in routes_df.columns:
        route_name_cols.append("route_short_name")
    if "route_long_name" in routes_df.columns:
        route_name_cols.append("route_long_name")
    route_type_df = routes_df.select(
        *[col(c).cast(StringType()) if c != "route_type" else col(c).cast(IntegerType())
          for c in route_name_cols]
    ).dropDuplicates(["route_id"])
    train_route_ids = [r.route_id for r in route_type_df.collect()]

    trips_df      = _read_gtfs_file(spark, feed_path, "trips.txt")
    stop_times_df = _read_gtfs_file(spark, feed_path, "stop_times.txt")
    stops_df      = _read_gtfs_file(spark, feed_path, "stops.txt")
    if trips_df is None or stop_times_df is None or stops_df is None:
        return None, None

    trips_df = trips_df.filter(col("route_id").cast(StringType()).isin(train_route_ids))
    if not trips_df.head(1):
        return None, None

    if station_id_mapping:
        import pandas as _pd
        mapping_df = spark.createDataFrame(
            _pd.DataFrame(list(station_id_mapping.items()),
                          columns=["old_stop_id", "canonical_stop_id"])
        )
        for attr_name, df_obj in (("stop_times_df", stop_times_df), ("stops_df", stops_df)):
            df_obj = (
                df_obj
                .join(broadcast(mapping_df),
                      df_obj["stop_id"].cast(StringType()) == mapping_df["old_stop_id"], "left")
                .withColumn("stop_id", coalesce(col("canonical_stop_id"), col("stop_id")))
                .drop("old_stop_id", "canonical_stop_id")
            )
            if attr_name == "stop_times_df":
                stop_times_df = df_obj
            else:
                stops_df = df_obj.dropDuplicates(["stop_id"])

    stop_times_train = stop_times_df.join(
        trips_df.select("trip_id", "route_id"), "trip_id", "inner"
    )
    stops_clean = stops_df.select(
        col("stop_id").cast(StringType()),
        col("stop_lat").cast(FloatType()).alias("lat"),
        col("stop_lon").cast(FloatType()).alias("lon"),
    ).filter(_is_valid_coord(col("lat"), col("lon")))

    pairs_with_dist = _build_stop_pairs_with_distance(
        stop_times_train, stops_clean, extra_cols=("route_id",)
    )
    logger.info("[TRAIN][TRANSFORM] %s/%s : tronçons calculés.", country_code, feed_id)
    if not pairs_with_dist.head(1):
        return None, None

    unique_routes = (
        pairs_with_dist
        .select(col("dep_stop_id").alias("dep_station_id"),
                col("arr_stop_id").alias("arr_station_id"), col("distance_km"))
        .groupBy("dep_station_id", "arr_station_id")
        .agg({"distance_km": "avg"})
        .withColumnRenamed("avg(distance_km)", "distance_km")
        .withColumn("is_night_train", lit(False))
    )

    pairs_with_type = pairs_with_dist.join(route_type_df, "route_id", "inner")
    name_cols = [c for c in ("route_short_name", "route_long_name") if c in pairs_with_type.columns]
    if name_cols:
        pairs_with_type = pairs_with_type.withColumn(
            "_route_names_upper", upper(concat_ws(" ", *[col(c) for c in name_cols]))
        )
        kw_cond = lit(False)
        for kw in TGV_KEYWORDS:
            kw_cond = kw_cond | col("_route_names_upper").contains(kw)
        pairs_with_type = pairs_with_type.withColumn(
            "vehicle_label",
            when(col("route_type").isin(list(TGV_ROUTE_TYPES)) | kw_cond, lit("TGV"))
            .otherwise(lit("Intercité"))
        ).drop("_route_names_upper")
    else:
        pairs_with_type = pairs_with_type.withColumn(
            "vehicle_label",
            when(col("route_type").isin(list(TGV_ROUTE_TYPES)), lit("TGV"))
            .otherwise(lit("Intercité"))
        )

    facts_with_co2 = (
        pairs_with_type
        .select(col("dep_stop_id").alias("dep_station_id"),
                col("arr_stop_id").alias("arr_station_id"),
                col("vehicle_label"), col("distance_km"))
        .groupBy("dep_station_id", "arr_station_id", "vehicle_label")
        .agg({"distance_km": "avg"})
        .withColumnRenamed("avg(distance_km)", "distance_km")
        .withColumn(
            "co2_kg_passenger",
            when(col("vehicle_label") == "TGV",
                 col("distance_km") * CO2_TGV_PER_100KM / 100.0)
            .otherwise(col("distance_km") * CO2_INTERCITE_PER_100KM / 100.0),
        )
    )
    return unique_routes, facts_with_co2


# ── TRAIN — LOAD ──────────────────────────────────────────────────────────────

def train_ensure_schema() -> None:
    _run_sql(f"CREATE SCHEMA IF NOT EXISTS {SCHEMA};")
    # Migration idempotente : source_id + élargissement station_id VARCHAR(50) -> VARCHAR(100)
    for stmt in [
        f"ALTER TABLE IF EXISTS {SCHEMA}.dim_station ADD COLUMN IF NOT EXISTS source_id INTEGER REFERENCES {SCHEMA}.dim_source(source_id);",
        f"ALTER TABLE IF EXISTS {SCHEMA}.dim_route_train ADD COLUMN IF NOT EXISTS source_id INTEGER REFERENCES {SCHEMA}.dim_source(source_id);",
        f"ALTER TABLE IF EXISTS {SCHEMA}.dim_route_avion ADD COLUMN IF NOT EXISTS source_id INTEGER REFERENCES {SCHEMA}.dim_source(source_id);",
        f"ALTER TABLE IF EXISTS {SCHEMA}.dim_vehicle_avion ADD COLUMN IF NOT EXISTS source_id INTEGER REFERENCES {SCHEMA}.dim_source(source_id);",
        # Colonne is_night_train (distinguer trains de jour vs nuit)
        f"ALTER TABLE IF EXISTS {SCHEMA}.dim_route_train ADD COLUMN IF NOT EXISTS is_night_train BOOLEAN NOT NULL DEFAULT FALSE;",
        # Élargir station_id à 100 chars (GTFS ids peuvent être longs)
        f"ALTER TABLE IF EXISTS {SCHEMA}.dim_station ALTER COLUMN station_id TYPE VARCHAR(100);",
        f"ALTER TABLE IF EXISTS {SCHEMA}.dim_route_train ALTER COLUMN dep_station_id TYPE VARCHAR(100);",
        f"ALTER TABLE IF EXISTS {SCHEMA}.dim_route_train ALTER COLUMN arr_station_id TYPE VARCHAR(100);",
        f"ALTER TABLE IF EXISTS {SCHEMA}.dim_route_avion ALTER COLUMN dep_station_id TYPE VARCHAR(100);",
        f"ALTER TABLE IF EXISTS {SCHEMA}.dim_route_avion ALTER COLUMN arr_station_id TYPE VARCHAR(100);",
    ]:
        _run_sql(stmt)


def train_truncate_tables() -> bool:
    logger.info("[TRAIN][LOAD] Vidage des tables train...")
    return _run_sql(
        f"TRUNCATE TABLE {', '.join(TABLES_TRAIN_TRUNCATE)} RESTART IDENTITY CASCADE;"
    )


def avion_truncate_tables() -> bool:
    logger.info("[AVION][LOAD] Vidage des tables avion...")
    return _run_sql(
        f"TRUNCATE TABLE {', '.join(TABLES_AVION_TRUNCATE)} RESTART IDENTITY CASCADE;"
    )


def _train_write_upsert(df, table_name: str, conflict_cols: str, update_cols: list = None) -> int:
    """INSERT … ON CONFLICT DO NOTHING (ou DO UPDATE) via psycopg2 execute_values (bulk).
    update_cols : liste de colonnes à mettre à jour sur conflit (ex: ['is_night_train']).
                  Si None → DO NOTHING.
    """
    if df is None:
        return 0
    rows = df.collect()
    if not rows:
        return 0
    columns = df.columns
    cols_str = ", ".join(f'"{c}"' for c in columns)
    if update_cols:
        update_clause = ", ".join(f'"{c}" = EXCLUDED."{c}"' for c in update_cols)
        sql = (
            f"INSERT INTO {table_name} ({cols_str}) VALUES %s "
            f"ON CONFLICT ({conflict_cols}) DO UPDATE SET {update_clause}"
        )
    else:
        sql = (
            f"INSERT INTO {table_name} ({cols_str}) VALUES %s "
            f"ON CONFLICT ({conflict_cols}) DO NOTHING"
        )
    # Conversion des valeurs Python → tuples psycopg2 (None conservé tel quel)
    data = [
        tuple(
            bool(row[c]) if isinstance(row[c], bool) else row[c]
            for c in columns
        )
        for row in rows
    ]
    try:
        from psycopg2.extras import execute_values
        conn = _get_pg_conn()
        conn.autocommit = False
        with conn.cursor() as cur:
            execute_values(cur, sql, data, page_size=1000)
        conn.commit()
        conn.close()
        logger.info("[TRAIN][LOAD] %s — %d lignes insérées.", table_name, len(data))
        return len(data)
    except Exception as e:
        logger.error("[TRAIN][LOAD] Erreur %s : %s", table_name, e)
        return 0


def train_seed_vehicles() -> None:
    """Insère les véhicules ferroviaires dans dim_vehicle_train."""
    logger.info("[TRAIN][LOAD] Seed dim_vehicle_train...")
    _run_sql(f"""
        INSERT INTO {SCHEMA}.dim_vehicle_train (label, co2_per_km, service_type) VALUES
            ('TGV',           0.0029, 'grande_vitesse'),
            ('Intercité',     0.0090, 'intercite'),
            ('Train de nuit', 0.0125, 'nuit')
        ON CONFLICT DO NOTHING;
    """)


def train_load_stations(spark, canonical_stations: list) -> int:
    """Insère les stations canoniques dans mart.dim_station (is_airport = FALSE).

    Bypass Spark intentionnel : createDataFrame sur 24k+ dicts lance des workers
    Python qui crashent sur Windows. Les données sont déjà en mémoire Python →
    insertion directe via psycopg2 par batchs de 1000 lignes.
    """
    if not canonical_stations:
        return 0
    columns = ["station_id", "station_name", "city", "country_code",
               "latitude", "longitude", "is_airport", "source_id"]
    cols_str = ", ".join(f'"{c}"' for c in columns)
    sql = (
        f"INSERT INTO {SCHEMA}.dim_station ({cols_str}) VALUES %s "
        f"ON CONFLICT (station_id) DO NOTHING"
    )
    data = [
        (
            str(row["station_id"]),
            row.get("station_name"),
            row.get("city"),
            row.get("country_code"),
            row.get("latitude"),
            row.get("longitude"),
            bool(row.get("is_airport", False)),
            int(row["source_id"]) if row.get("source_id") is not None else None,
        )
        for row in canonical_stations
    ]
    try:
        from psycopg2.extras import execute_values
        conn = _get_pg_conn()
        conn.autocommit = False
        with conn.cursor() as cur:
            execute_values(cur, sql, data, page_size=1000)
        conn.commit()
        conn.close()
        logger.info("[TRAIN][LOAD] %s.dim_station — %d lignes insérées.", SCHEMA, len(data))
        return len(data)
    except Exception as e:
        logger.error("[TRAIN][LOAD] dim_station — erreur : %s", e)
        return 0


def train_insert_facts(facts_df) -> None:
    """Insère les faits ferroviaires dans fact_emission (résolution FK via PostgreSQL)."""
    rows = facts_df.collect()
    if not rows:
        return
    logger.info("[TRAIN][LOAD] Insertion de %d faits dans fact_emission...", len(rows))
    try:
        with tempfile.NamedTemporaryFile(mode="w", suffix=".sql", delete=False, encoding="utf-8") as f:
            tmp = f.name
            for row in rows:
                dep = str(row["dep_station_id"]).replace("'", "''")
                arr = str(row["arr_station_id"]).replace("'", "''")
                veh = str(row["vehicle_label"]).replace("'", "''")
                co2 = row["co2_kg_passenger"]
                f.write(
                    f"INSERT INTO {SCHEMA}.fact_emission "
                    f"(transport_mode, route_train_id, vehicle_train_id, co2_kg_passenger)\n"
                    f"SELECT 'train', r.route_train_id, v.vehicle_train_id, {co2}\n"
                    f"FROM {SCHEMA}.dim_route_train r, {SCHEMA}.dim_vehicle_train v\n"
                    f"WHERE r.dep_station_id = '{dep}'\n"
                    f"  AND r.arr_station_id = '{arr}'\n"
                    f"  AND v.label = '{veh}'\n"
                    f"ON CONFLICT (route_train_id, vehicle_train_id) DO NOTHING;\n"
                )
        ok = _run_sql_file(tmp)
        os.remove(tmp)
        if ok:
            logger.info("[TRAIN][LOAD] fact_emission — %d faits traités.", len(rows))
    except Exception as e:
        logger.error("[TRAIN][LOAD] fact_emission — erreur : %s", e)


def train_process_feed(spark, feed_info: dict, global_mapping: dict, source_id: "int | None" = None) -> int:
    """Traite un feed GTFS complet (routes + faits). Retourne le nombre de routes insérées."""
    country_code = feed_info["country_code"]
    feed_id      = feed_info["feed_id"]
    feed_path    = feed_info["path"]
    logger.info("[TRAIN][LOAD] Traitement feed %s/%s...", country_code, feed_id)
    try:
        feed_mapping = train_build_feed_stop_mapping(spark, feed_path, global_mapping)
        logger.info("[TRAIN][LOAD]   %d stop_ids remappés.", len(feed_mapping))
        routes_df, facts_df = train_process_routes_and_facts(
            spark, feed_path, feed_id, country_code, feed_mapping
        )
        if routes_df is None:
            logger.info("[TRAIN][LOAD]   Aucune route valide.")
            return 0
        if source_id is not None:
            routes_df = routes_df.withColumn("source_id", lit(source_id).cast(IntegerType()))
        n = _train_write_upsert(routes_df, f"{SCHEMA}.dim_route_train", "dep_station_id, arr_station_id")
        if facts_df is not None:
            train_insert_facts(facts_df)
        return n
    except Exception as e:
        logger.error("[TRAIN][LOAD] Erreur feed %s/%s : %s", country_code, feed_id, e)
        return 0


def train_process_backontrack(spark, global_mapping: dict, source_id: "int | None" = None) -> int:
    """Traite les routes Back on Track (trains de nuit). Retourne le nombre de routes insérées."""
    trip_stop_path = os.path.join(BACKONTRACK_DIR, "back_on_track_trip_stop.csv")
    stops_path     = os.path.join(BACKONTRACK_DIR, "back_on_track_stops.csv")
    if not os.path.exists(trip_stop_path) or not os.path.exists(stops_path):
        logger.warning("[TRAIN][LOAD] Back on Track : fichiers manquants, ignoré.")
        return 0
    logger.info("[TRAIN][LOAD] Back on Track (trains de nuit)...")
    try:
        trip_stop_df = spark.read.csv(trip_stop_path, header=True, inferSchema=True)
        stops_df     = (
            spark.read.csv(stops_path, header=True, inferSchema=True)
            .withColumn("lat", col("station_lat").cast(FloatType()))
            .withColumn("lon", col("station_long").cast(FloatType()))
        )
        stops_df = stops_df.filter(_is_valid_coord(col("lat"), col("lon")))

        if global_mapping:
            import pandas as _pd
            mapping_df = spark.createDataFrame(
                _pd.DataFrame(list(global_mapping.items()),
                              columns=["old_stop_id", "canonical_stop_id"])
            )
            for attr_name, df_obj in (("trip_stop_df", trip_stop_df), ("stops_df", stops_df)):
                df_obj = (
                    df_obj
                    .join(broadcast(mapping_df),
                          df_obj["stop_id"].cast(StringType()) == mapping_df["old_stop_id"], "left")
                    .withColumn("stop_id", coalesce(col("canonical_stop_id"), col("stop_id")))
                    .drop("old_stop_id", "canonical_stop_id")
                )
                if attr_name == "trip_stop_df":
                    trip_stop_df = df_obj
                else:
                    stops_df = df_obj.dropDuplicates(["stop_id"])

        stops_clean = stops_df.select(col("stop_id").cast(StringType()), col("lat"), col("lon"))
        pairs_with_dist = _build_stop_pairs_with_distance(trip_stop_df, stops_clean)
        if not pairs_with_dist.head(1):
            return 0

        unique_routes = (
            pairs_with_dist
            .select(col("dep_stop_id").alias("dep_station_id"),
                    col("arr_stop_id").alias("arr_station_id"), col("distance_km"))
            .groupBy("dep_station_id", "arr_station_id")
            .agg({"distance_km": "avg"})
            .withColumnRenamed("avg(distance_km)", "distance_km")
            .withColumn("is_night_train", lit(True))
        )
        if source_id is not None:
            unique_routes = unique_routes.withColumn("source_id", lit(source_id).cast(IntegerType()))
        # DO UPDATE is_night_train : si la route existe déjà (depuis GTFS), on la marque comme nuit aussi
        n = _train_write_upsert(unique_routes, f"{SCHEMA}.dim_route_train",
                                "dep_station_id, arr_station_id",
                                update_cols=["is_night_train"])

        facts_with_co2 = (
            pairs_with_dist
            .select(col("dep_stop_id").alias("dep_station_id"),
                    col("arr_stop_id").alias("arr_station_id"), col("distance_km"))
            .groupBy("dep_station_id", "arr_station_id")
            .agg({"distance_km": "avg"})
            .withColumnRenamed("avg(distance_km)", "distance_km")
            .withColumn("vehicle_label", lit("Train de nuit"))
            .withColumn("co2_kg_passenger",
                        col("distance_km") * CO2_NIGHTTRAIN_PER_100KM / 100.0)
        )
        train_insert_facts(facts_with_co2)
        logger.info("[TRAIN][LOAD] Back on Track : %d routes de nuit insérées.", n)
        return n
    except Exception as e:
        logger.error("[TRAIN][LOAD] Back on Track — erreur : %s", e)
        return 0


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║                          SECTION  AVION                                  ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

# ── AVION — CONFIGURATION ─────────────────────────────────────────────────────

AIRPORTS_CSV   = "./airports.csv"
EASA_FILE      = "./EASA_CO2_Database__web__-Issue11.xlsx"
EASA_SHEET     = "EASA CO2DB (web)"
OPDI_CACHE_DIR = "./opdi_cache"   # Cache disque des parquets mensuels

BASE_URL   = "https://www.eurocontrol.int/performance/data/download/OPDI/v002/flight_list/flight_list_"
START_DATE = "202501"
END_DATE   = "202512"

# Correspondance désignation EASA → code ICAO type (Source : ICAO Doc 8643)
EASA_TO_ICAO: dict = {
    "A319-151N":  "A19N", "A319-153N":  "A19N", "A319-171N":  "A19N",
    "A320-251N":  "A20N", "A320-252N":  "A20N", "A320-253N":  "A20N",
    "A320-271N":  "A20N", "A320-272N":  "A20N", "A320-273N":  "A20N",
    "A321-251NX": "A21N", "A321-252NX": "A21N", "A321-253NX": "A21N",
    "A321-253NY": "A21N", "A321-271NX": "A21N", "A321-271NY": "A21N",
    "A321-272NX": "A21N",
    "A330-841":   "A338", "A330-941":   "A339",
    "A350-941":   "A359", "A350-1041":  "A35K",
    "ATR 72-212A": "AT72",
    "BD-500-1A10": "BCS1", "BD-500-1A11": "BCS3",
    "Falcon 7X":  "F7X",
}


def _avion_generate_urls() -> list:
    start_dt = datetime.strptime(START_DATE, "%Y%m")
    end_dt   = datetime.strptime(END_DATE,   "%Y%m")
    urls, current = [], start_dt
    while current <= end_dt:
        urls.append(f"{BASE_URL}{current.strftime('%Y%m')}.parquet")
        current += relativedelta(months=1)
    return urls


def _avion_classify_service(mtom_kg: "int | None") -> str:
    """Classe un avion selon sa MTOM (kg) en catégorie de service."""
    if mtom_kg is None:
        return "inconnu"
    if mtom_kg < 40_000:
        return "regional"
    if mtom_kg < 100_000:
        return "court_moyen_courrier"
    return "long_courrier"


# ── AVION — EXTRACT ───────────────────────────────────────────────────────────

def avion_extract_easa(filepath: str, sheet: str) -> "dict | None":
    """
    Lit le fichier Excel EASA CO2 Database avec openpyxl (PySpark ne lit pas Excel).
    Agrège les valeurs CO2 par désignation de type d'avion.
    Colonnes utilisées (index 0-based) :
      [2]  Aeroplane Type Certificate Holder → manufacturer
      [3]  Aeroplane Type Designation        → label
      [8]  Aeroplane MTOM (kg)               → mtom_kg
      [14] Number of Engines                 → num_engines
      [26] CO2 Metric Value (kg/km)          → co2_per_km
    """
    logger.info("[AVION][EXTRACT] EASA : %s (feuille : %s)", filepath, sheet)
    try:
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Fichier EASA introuvable : {filepath}")
        wb   = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            raise ValueError(f"Feuille '{sheet}' absente. Disponibles : {wb.sheetnames}")
        rows  = list(wb[sheet].iter_rows(values_only=True))[3:]
        types: dict = defaultdict(
            lambda: {"co2_values": [], "mtom": None, "engines": None, "holder": None}
        )
        for r in rows:
            designation = r[3]
            co2         = r[26]
            if not designation or co2 is None:
                continue
            key = str(designation).strip()
            try:
                types[key]["co2_values"].append(float(co2))
                if types[key]["mtom"]    is None and r[8]  is not None:
                    types[key]["mtom"]    = int(r[8])
                if types[key]["engines"] is None and r[14] is not None:
                    types[key]["engines"] = int(r[14])
                if types[key]["holder"]  is None and r[2]  is not None:
                    types[key]["holder"]  = str(r[2]).strip()
            except (ValueError, TypeError) as e:
                logger.warning("[AVION][EXTRACT] Ligne EASA ignorée (%s) : %s", key, e)
        if not types:
            raise ValueError("Aucun type d'avion extrait.")
        logger.info("[AVION][EXTRACT] %d types d'avions extraits.", len(types))
        return dict(types)
    except Exception as e:
        logger.error("[AVION][EXTRACT] EASA — erreur : %s", e)
    return None


def avion_load_airports_df(csv_path: str, spark) -> "object | None":
    """
    Charge airports.csv comme Spark DataFrame.
    Colonnes résultantes : airport_id, lat, lon, airport_name, country_code.
    Retourne None en cas d'erreur.
    """
    logger.info("[AVION][EXTRACT] Aéroports : %s", csv_path)
    try:
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"Fichier introuvable : {csv_path}")
        df = spark.read.csv(csv_path, header=True, inferSchema=True)
        required = {"ident", "latitude", "longitude", "name", "country"}
        missing  = required - set(df.columns)
        if missing:
            raise ValueError(f"Colonnes manquantes dans airports.csv : {missing}")
        result = (
            df.select(
                col("ident").cast(StringType()).alias("airport_id"),
                col("latitude").cast(FloatType()).alias("lat"),
                col("longitude").cast(FloatType()).alias("lon"),
                col("name").cast(StringType()).alias("airport_name"),
                col("country").cast(StringType()).alias("country_code"),
                col("municipality").cast(StringType()).alias("city"),
            )
            .filter(col("airport_id").isNotNull())
            .filter(col("lat").isNotNull() & col("lon").isNotNull())
            .dropDuplicates(["airport_id"])
        )
        logger.info("[AVION][EXTRACT] %d aéroports chargés.", result.count())
        return result
    except Exception as e:
        logger.error("[AVION][EXTRACT] Aéroports — erreur : %s", e)
    return None


def avion_download_to_cache(url: str, month: str) -> "str | None":
    """
    Télécharge le parquet mensuel Eurocontrol OPDI dans OPDI_CACHE_DIR.
    Si le fichier est déjà présent (cache hit), retourne directement le chemin.
    Aucune donnée n'est chargée en mémoire ici — la lecture est faite par PySpark.
    """
    os.makedirs(OPDI_CACHE_DIR, exist_ok=True)
    path = os.path.join(OPDI_CACHE_DIR, f"flight_list_{month}.parquet")
    if os.path.exists(path):
        logger.info("[AVION][EXTRACT] Cache hit : %s", month)
        return path
    logger.info("[AVION][EXTRACT] Téléchargement : %s", month)
    try:
        response = requests.get(url, timeout=180)
        response.raise_for_status()
        with open(path, "wb") as f:
            f.write(response.content)
        size_mb = len(response.content) / 1_000_000
        logger.info("[AVION][EXTRACT] %s téléchargé (%.1f Mo).", month, size_mb)
        return path
    except requests.exceptions.HTTPError as e:
        logger.warning("[AVION][EXTRACT] HTTP %s pour %s — ignoré.", e.response.status_code, month)
    except requests.exceptions.ConnectionError:
        logger.warning("[AVION][EXTRACT] Connexion échouée pour %s — ignoré.", month)
    except requests.exceptions.Timeout:
        logger.warning("[AVION][EXTRACT] Timeout pour %s — ignoré.", month)
    except Exception as e:
        logger.warning("[AVION][EXTRACT] Erreur %s : %s", month, e)
    # Supprimer le fichier partiel si présent
    if os.path.exists(path):
        os.remove(path)
    return None


# ── AVION — TRANSFORM ─────────────────────────────────────────────────────────

def avion_transform_easa(types: dict) -> list:
    """
    Construit les tuples prêts à l'insertion pour dim_vehicle_avion.
    CO2 = moyenne des variants certifiés du même type d'avion.
    """
    rows = []
    for designation, data in sorted(types.items()):
        try:
            if not data["co2_values"]:
                logger.warning("[AVION][TRANSFORM] %s : aucune valeur CO2 — ignoré.", designation)
                continue
            avg_co2      = round(mean(data["co2_values"]), 4)
            mtom         = data["mtom"]
            service_type = _avion_classify_service(mtom)
            icao_tc      = EASA_TO_ICAO.get(designation)
            rows.append((designation, avg_co2, service_type, icao_tc, mtom,
                         data["engines"], data["holder"]))
            logger.info(
                "[AVION][TRANSFORM]  %-35s  CO2=%.4f kg/km  service=%-22s  ICAO=%s",
                designation, avg_co2, service_type, icao_tc or "-",
            )
        except Exception as e:
            logger.warning("[AVION][TRANSFORM] %s : erreur (%s) — ignoré.", designation, e)
    logger.info("[AVION][TRANSFORM] %d types d'avions prêts à l'insertion.", len(rows))
    return rows


def avion_transform_all_flights(spark, airports_df) -> tuple:
    """
    Lit tous les parquets du cache (OPDI_CACHE_DIR) en une seule passe PySpark.
    Joint avec airports_df pour les coordonnées GPS.
    Calcule la distance haversine (colonne Spark partagée avec train).
    Retourne :
      - routes_df   : (adep, ades, distance_km) — distance moyenne par route
      - tc_counts_df: (adep, ades, typecode, cnt) — comptage typecodes par route
    """
    cached_files = [
        os.path.join(OPDI_CACHE_DIR, f)
        for f in os.listdir(OPDI_CACHE_DIR)
        if f.startswith("flight_list_") and f.endswith(".parquet")
    ]
    if not cached_files:
        logger.warning("[AVION][TRANSFORM] Aucun fichier parquet en cache.")
        return None, None

    logger.info("[AVION][TRANSFORM] Lecture de %d fichiers parquet via PySpark...", len(cached_files))
    # Schéma minimal — évite le crash sur la colonne uint64 (id) non supportée par Spark
    # La projection de colonnes permet d'ignorer les colonnes problématiques
    from pyspark.sql.types import StructType, StructField
    minimal_schema = StructType([
        StructField("adep",     StringType(), True),
        StructField("ades",     StringType(), True),
        StructField("typecode", StringType(), True),
    ])
    # Chemins explicites (pas de glob Hadoop — hang sur Windows sans winutils)
    all_flights = (
        spark.read.schema(minimal_schema).parquet(*cached_files)
        .filter(
            col("adep").isNotNull() & (col("adep") != "")
            & col("ades").isNotNull() & (col("ades") != "")
        )
    )
    logger.info("[AVION][TRANSFORM] Parquet charge et filtre (adep+ades non nuls).")

    # Jointure avec les aéroports pour les coordonnées GPS (broadcast — table petite)
    airports_broadcast = broadcast(airports_df)
    flights_with_coords = (
        all_flights
        .join(
            airports_broadcast.select(
                col("airport_id").alias("adep"),
                col("lat").alias("dep_lat"),
                col("lon").alias("dep_lon"),
            ),
            "adep", "inner",   # inner : on ne garde que les vols avec aéroports connus
        )
        .join(
            airports_broadcast.select(
                col("airport_id").alias("ades"),
                col("lat").alias("arr_lat"),
                col("lon").alias("arr_lon"),
            ),
            "ades", "inner",
        )
    )

    # Calcul de la distance haversine (fonction partagée avec la section train)
    flights_with_dist = flights_with_coords.withColumn(
        "distance_km",
        _haversine_col(col("dep_lat"), col("dep_lon"), col("arr_lat"), col("arr_lon")),
    )

    # Agrégation : distance moyenne par route (adep, ades)
    routes_df = (
        flights_with_dist
        .groupBy("adep", "ades")
        .agg(sp_avg("distance_km").alias("distance_km"))
    )

    # Comptage des typecodes par route (pour trouver le dominant)
    tc_counts_df = (
        flights_with_dist
        .filter(col("typecode").isNotNull() & (col("typecode") != ""))
        .groupBy("adep", "ades", "typecode")
        .agg(sp_count("*").alias("cnt"))
    )

    logger.info("[AVION][TRANSFORM] Agregation routes et typecodes terminee.")
    return routes_df, tc_counts_df


def avion_consolidate_routes(routes_df, tc_counts_df) -> "object":
    """
    Sélectionne le typecode dominant par route (le plus fréquent sur toute l'année).
    Joint avec routes_df pour produire le DataFrame final :
      colonnes : (adep, ades, distance_km, dominant_typecode)
    """
    # Trouver le typecode avec le plus grand comptage par route
    w = Window.partitionBy("adep", "ades").orderBy(col("cnt").desc())
    dominant_tc = (
        tc_counts_df
        .withColumn("rn", row_number().over(w))
        .filter(col("rn") == 1)
        .select("adep", "ades", col("typecode").alias("dominant_typecode"))
    )
    final_routes = routes_df.join(dominant_tc, ["adep", "ades"], "left")
    with_tc = final_routes.filter(col("dominant_typecode").isNotNull()).count()
    total   = final_routes.count()
    logger.info(
        "[AVION][TRANSFORM] %d routes finales | %d avec typecode dominant.",
        total, with_tc,
    )
    return final_routes


# ── AVION — LOAD ──────────────────────────────────────────────────────────────

def avion_load_vehicle_psql(easa_rows: list, source_id: "int | None" = None) -> bool:
    """
    Migration idempotente des colonnes de dim_vehicle_avion.
    Upsert des types d'avions EASA via psql (ON CONFLICT sur label).
    """
    logger.info("[AVION][LOAD] dim_vehicle_avion : %d type(s)", len(easa_rows))
    for sql in [
        "ALTER TABLE mart.dim_vehicle_avion ADD COLUMN IF NOT EXISTS icao_typecode VARCHAR(10);",
        "ALTER TABLE mart.dim_vehicle_avion ADD COLUMN IF NOT EXISTS mtom_kg INTEGER;",
        "ALTER TABLE mart.dim_vehicle_avion ADD COLUMN IF NOT EXISTS num_engines SMALLINT;",
        "ALTER TABLE mart.dim_vehicle_avion ADD COLUMN IF NOT EXISTS manufacturer VARCHAR(150);",
        """DO $$ BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'dim_vehicle_avion_label_key')
            THEN ALTER TABLE mart.dim_vehicle_avion
                 ADD CONSTRAINT dim_vehicle_avion_label_key UNIQUE (label);
            END IF;
        END $$;""",
    ]:
        _run_sql(sql)

    try:
        def esc(v):
            return "NULL" if v is None else f"'{str(v).replace(chr(39), chr(39)*2)}'"

        def num(v):
            return "NULL" if v is None else str(v)

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".sql", delete=False, encoding="utf-8"
        ) as f:
            tmp = f.name
            for designation, avg_co2, service_type, icao_tc, mtom, engines, holder in easa_rows:
                f.write(
                    f"INSERT INTO mart.dim_vehicle_avion "
                    f"(label, co2_per_km, service_type, icao_typecode, mtom_kg, num_engines, manufacturer, source_id) "
                    f"VALUES ({esc(designation)}, {avg_co2}, {esc(service_type)}, {esc(icao_tc)}, "
                    f"{num(mtom)}, {num(engines)}, {esc(holder)}, {num(source_id)}) "
                    f"ON CONFLICT (label) DO UPDATE "
                    f"SET co2_per_km=EXCLUDED.co2_per_km, service_type=EXCLUDED.service_type, "
                    f"icao_typecode=EXCLUDED.icao_typecode, mtom_kg=EXCLUDED.mtom_kg, "
                    f"num_engines=EXCLUDED.num_engines, manufacturer=EXCLUDED.manufacturer, "
                    f"source_id=EXCLUDED.source_id;\n"
                )
        ok = _run_sql_file(tmp)
        os.remove(tmp)
        if ok:
            logger.info("[AVION][LOAD] dim_vehicle_avion — %d types insérés/mis à jour.", len(easa_rows))
        return ok
    except Exception as e:
        logger.error("[AVION][LOAD] dim_vehicle_avion — erreur : %s", e)
        return False


def avion_load_stations_psql(final_routes_df, airports_df, source_id: "int | None" = None) -> int:
    """
    Extrait les aéroports uniques (adep + ades) depuis les routes finales,
    joint avec airports_df pour les métadonnées GPS,
    insère dans mart.dim_station (is_airport = TRUE) via psql.
    """
    logger.info("[AVION][LOAD] dim_station (aéroports)...")
    try:
        airports_used = (
            final_routes_df.select(col("adep").alias("airport_id"))
            .union(final_routes_df.select(col("ades").alias("airport_id")))
            .distinct()
            .join(airports_df, "airport_id", "inner")
        )
        rows = airports_used.collect()
        if not rows:
            logger.warning("[AVION][LOAD] Aucun aéroport à insérer.")
            return 0

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".sql", delete=False, encoding="utf-8"
        ) as f:
            tmp = f.name
            for row in rows:
                aid  = str(row["airport_id"]).replace("'", "''")
                name = str(row["airport_name"] or row["airport_id"]).replace("'", "''")
                city = str(row["city"] or row["airport_name"] or "").replace("'", "''")
                cc   = str(row["country_code"] or "XX")[:2]
                lat  = row["lat"] if row["lat"] is not None else "NULL"
                lon  = row["lon"] if row["lon"] is not None else "NULL"
                src = source_id if source_id is not None else "NULL"
                f.write(
                    f"INSERT INTO mart.dim_station "
                    f"(station_id, station_name, city, country_code, is_airport, latitude, longitude, source_id) "
                    f"VALUES ('{aid}', '{name}', '{city}', '{cc}', TRUE, {lat}, {lon}, {src}) "
                    f"ON CONFLICT (station_id) DO UPDATE "
                    f"SET station_name=EXCLUDED.station_name, city=EXCLUDED.city, "
                    f"latitude=EXCLUDED.latitude, longitude=EXCLUDED.longitude, "
                    f"source_id=EXCLUDED.source_id;\n"
                )
        ok = _run_sql_file(tmp)
        os.remove(tmp)
        if ok:
            logger.info("[AVION][LOAD] dim_station — %d aéroports insérés.", len(rows))
            return len(rows)
        return 0
    except Exception as e:
        logger.error("[AVION][LOAD] dim_station — erreur : %s", e)
        return 0


def avion_load_routes_psql(final_routes_df, source_id: "int | None" = None) -> int:
    """
    Migration des colonnes dominant_typecode et co2_total_kg (idempotente).
    Insère les routes dans mart.dim_route_avion via psql (ON CONFLICT upsert).
    """
    logger.info("[AVION][LOAD] dim_route_avion...")
    for sql in [
        "ALTER TABLE mart.dim_route_avion ADD COLUMN IF NOT EXISTS dominant_typecode VARCHAR(50);",
        "ALTER TABLE mart.dim_route_avion ALTER COLUMN dominant_typecode TYPE VARCHAR(50);",
        "ALTER TABLE mart.dim_route_avion ADD COLUMN IF NOT EXISTS co2_total_kg NUMERIC(12,3);",
    ]:
        _run_sql(sql)

    try:
        rows = final_routes_df.collect()
        if not rows:
            logger.warning("[AVION][LOAD] Aucune route à insérer.")
            return 0

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".sql", delete=False, encoding="utf-8"
        ) as f:
            tmp = f.name
            for row in rows:
                adep = str(row["adep"]).replace("'", "''")
                ades = str(row["ades"]).replace("'", "''")
                dist = row["distance_km"] if row["distance_km"] is not None else "NULL"
                tc   = (
                    f"'{str(row['dominant_typecode'])[:50].replace(chr(39), chr(39)*2)}'"
                    if row["dominant_typecode"] else "NULL"
                )
                src = source_id if source_id is not None else "NULL"
                f.write(
                    f"INSERT INTO mart.dim_route_avion "
                    f"(dep_station_id, arr_station_id, distance_km, dominant_typecode, source_id) "
                    f"VALUES ('{adep}', '{ades}', {dist}, {tc}, {src}) "
                    f"ON CONFLICT (dep_station_id, arr_station_id) DO UPDATE "
                    f"SET distance_km=EXCLUDED.distance_km, "
                    f"dominant_typecode=EXCLUDED.dominant_typecode, "
                    f"source_id=EXCLUDED.source_id;\n"
                )
        ok = _run_sql_file(tmp)
        os.remove(tmp)
        if ok:
            logger.info("[AVION][LOAD] dim_route_avion — %d routes insérées.", len(rows))
            return len(rows)
        return 0
    except Exception as e:
        logger.error("[AVION][LOAD] dim_route_avion — erreur : %s", e)
        return 0


def avion_update_co2_psql() -> None:
    """
    Calcule co2_total_kg en deux passes SQL (psql) :
      1. JOIN exact EASA : distance_km × co2_per_km du typecode dominant
      2. Fallback catégorie ICAO : distance_km × AVG(co2_per_km) par tranche
         < 500 km → regional | < 4000 km → court_moyen_courrier | ≥ 4000 km → long_courrier
    """
    logger.info("[AVION][LOAD] Calcul co2_total_kg...")
    ok1 = _run_sql("""
        UPDATE mart.dim_route_avion r
        SET co2_total_kg = ROUND(r.distance_km::NUMERIC * v.co2_per_km, 3)
        FROM mart.dim_vehicle_avion v
        WHERE v.icao_typecode = r.dominant_typecode
          AND r.distance_km IS NOT NULL;
    """)
    if ok1:
        logger.info("[AVION][LOAD] co2_total_kg (exact EASA) : mise à jour OK.")

    ok2 = _run_sql("""
        UPDATE mart.dim_route_avion r
        SET co2_total_kg = ROUND(
            r.distance_km::NUMERIC * (
                SELECT AVG(v.co2_per_km)
                FROM mart.dim_vehicle_avion v
                WHERE v.service_type = CASE
                    WHEN r.distance_km <  500  THEN 'regional'
                    WHEN r.distance_km < 4000  THEN 'court_moyen_courrier'
                    ELSE                             'long_courrier'
                END
            ), 3
        )
        WHERE r.co2_total_kg IS NULL
          AND r.distance_km IS NOT NULL;
    """)
    if ok2:
        logger.info("[AVION][LOAD] co2_total_kg (fallback catégorie ICAO) : mise à jour OK.")


def avion_insert_facts_psql() -> int:
    """
    Insère les faits avion dans fact_emission après calcul co2_total_kg.
    co2_kg_passenger = co2_total_kg / nb_passagers_moyen selon catégorie de service.
      - regional           : 70 passagers
      - court_moyen_courrier: 150 passagers
      - long_courrier       : 280 passagers
      - inconnu (NULL)      : 150 passagers par défaut
    """
    logger.info("[AVION][LOAD] Insertion des faits avion dans fact_emission...")
    ok = _run_sql(f"""
        INSERT INTO {SCHEMA}.fact_emission
            (transport_mode, route_avion_id, vehicle_avion_id, co2_kg_passenger)
        SELECT
            'avion',
            r.route_avion_id,
            v.vehicle_avion_id,
            ROUND(
                r.co2_total_kg / CASE
                    WHEN v.service_type = 'regional'              THEN 70.0
                    WHEN v.service_type = 'court_moyen_courrier'  THEN 150.0
                    WHEN v.service_type = 'long_courrier'         THEN 280.0
                    ELSE 150.0
                END
            , 3)
        FROM {SCHEMA}.dim_route_avion r
        LEFT JOIN (
            SELECT DISTINCT ON (icao_typecode)
                vehicle_avion_id, service_type, icao_typecode
            FROM {SCHEMA}.dim_vehicle_avion
            WHERE icao_typecode IS NOT NULL
            ORDER BY icao_typecode, vehicle_avion_id
        ) v ON v.icao_typecode = r.dominant_typecode
        WHERE r.co2_total_kg IS NOT NULL
        ON CONFLICT (route_avion_id, vehicle_avion_id) DO NOTHING;
    """)
    if ok:
        raw = _query_sql(f"SELECT COUNT(*) FROM {SCHEMA}.fact_emission WHERE transport_mode = 'avion';")
        n = int(raw.strip()) if raw and raw.strip().isdigit() else 0
        logger.info("[AVION][LOAD] fact_emission — %d faits avion insérés.", n)
        return n
    logger.error("[AVION][LOAD] fact_emission avion — échec de l'insertion.")
    return 0


# ═══════════════════════════════════════════════════════════════════════════════
# ORCHESTRATION
# ═══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    logger.info("=== Pipeline ETL RailCarbon — début (train + avion, PySpark) ===")

    # Métriques globales cumulées (train + avion) pour mart.etl_run_log
    metrics: dict = {
        "mois_traites":     0,
        "airports_charges": 0,   # stations train + aéroports avion
        "routes_chargees":  0,   # routes train + routes avion
        "co2_exact":        0,
        "co2_fallback":     0,
        "erreurs":          0,
    }

    # Démarrage du suivi dans mart.etl_run_log
    run_id = run_start()

    # ── Seeding des sources statiques (EASA, OPDI, OurAirports, Back on Track) ─
    source_ids = source_seed_static()
    logger.info("[SOURCE] Sources statiques chargées : %s", list(source_ids.keys()))

    # ── Construction du lookup feed_id → url depuis feeds_v2.csv ──────────────
    feed_url_lookup: dict = {}
    try:
        import csv
        with open(FEEDS_CSV, encoding="utf-8") as fcsv:
            for row in csv.DictReader(fcsv):
                fid = row.get("id", "").strip()
                url = row.get("urls.latest", "").strip()
                if fid and url:
                    feed_url_lookup[fid] = url
    except Exception as _e:
        logger.warning("[SOURCE] Impossible de lire les URLs feeds_v2.csv : %s", _e)

    # ── Session Spark unique (partagée train + avion) ──────────────────────────
    # Hadoop 3.x appelle Subject.getSubject() supprime en Java 21+ : crash.
    # Solution : forcer JAVA_HOME vers Java 17 LTS (compatible Hadoop 3.x).
    import glob as _glob
    _java17_search = [
        r"C:\Program Files\Eclipse Adoptium\jdk-17*",
        r"C:\Program Files\Microsoft\jdk-17*",
        r"C:\Program Files\Java\jdk-17*",
        r"C:\Program Files\BellSoft\LibericaJDK-17*",
    ]
    _java17_found = None
    for _pattern in _java17_search:
        _matches = sorted(_glob.glob(_pattern), reverse=True)  # plus récent en premier
        if _matches:
            _java17_found = _matches[0]
            break
    if _java17_found:
        os.environ["JAVA_HOME"] = _java17_found
        os.environ["PATH"] = os.path.join(_java17_found, "bin") + os.pathsep + os.environ.get("PATH", "")
        logger.info("[SPARK] JAVA_HOME force vers Java 17 : %s", _java17_found)
    else:
        logger.warning("[SPARK] Java 17 non trouve. Installez Temurin 17 : https://adoptium.net/")
    _jars_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "jars")
    os.makedirs(_jars_dir, exist_ok=True)
    _pg_jar = os.path.join(_jars_dir, "postgresql-42.7.1.jar")
    if not os.path.exists(_pg_jar):
        logger.info("[SPARK] Telechargement JAR PostgreSQL JDBC (premiere execution)...")
        urllib.request.urlretrieve(
            "https://repo1.maven.org/maven2/org/postgresql/postgresql/42.7.1/postgresql-42.7.1.jar",
            _pg_jar,
        )
        logger.info("[SPARK] JAR telecharge : %s", _pg_jar)
    else:
        logger.info("[SPARK] JAR PostgreSQL en cache : %s", _pg_jar)

    spark = (
        SparkSession.builder
        .appName("RailCarbon ETL")
        .config("spark.driver.extraClassPath", _pg_jar)
        .config("spark.executor.extraClassPath", _pg_jar)
        .config("spark.driver.memory", "4g")           # 8g inutile en local mode
        .config("spark.executor.memory", "2g")         # driver=executor en local
        .config("spark.sql.shuffle.partitions", "8")   # défaut 200 = gaspillage en local
        .config("spark.python.worker.reuse", "true")   # évite de respawn des workers à chaque tâche
        .config("spark.sql.execution.arrow.pyspark.enabled", "true")  # pandas↔Spark via Arrow (pas Py4J)
        .getOrCreate()
    )

    try:
        # ══════════════════════════════════════════════════════════════════════
        # ETL TRAIN
        # ══════════════════════════════════════════════════════════════════════
        logger.info("--- ETL TRAIN -----------------------------------------------------------")

        train_ensure_schema()
        if not train_truncate_tables():
            logger.error("[TRAIN] Echec du vidage des tables — ETL train abandonné.")
            metrics["erreurs"] += 1
        else:
            train_seed_vehicles()

            # ── TRAIN — EXTRACT ────────────────────────────────────────────────
            logger.info("--- TRAIN - EXTRACT ---------------------------------------------")
            train_download_all_gtfs(spark)
            feeds = train_list_gtfs_feeds(GTFS_DIR)
            logger.info("[TRAIN][EXTRACT] %d feeds GTFS à traiter.", len(feeds))

            all_train_stations = []
            feeds_with_trains  = []

            for feed in feeds:
                try:
                    stop_ids = train_get_stop_ids(spark, feed["path"])
                    if not stop_ids:
                        continue
                    # Upsert source GTFS pour ce feed précis
                    feed_url = feed_url_lookup.get(str(feed["feed_id"]), "")
                    feed["source_id"] = source_upsert_gtfs_feed(
                        feed["feed_id"], feed["country_code"], feed_url
                    )
                    feeds_with_trains.append(feed)
                    stations = train_collect_stations_from_feed(
                        spark, feed["path"], feed["country_code"], stop_ids,
                        source_id=feed["source_id"]
                    )
                    all_train_stations.extend(stations)
                    logger.info("[TRAIN][EXTRACT]   %s : %d stations", feed["feed_id"], len(stations))
                    spark.catalog.clearCache()
                except Exception as e:
                    logger.error("[TRAIN][EXTRACT] Erreur collecte %s : %s", feed["feed_id"], e)
                    metrics["erreurs"] += 1

            try:
                bot_stations = train_collect_stations_from_backontrack(
                    spark, source_id=source_ids.get("Back on Track")
                )
                all_train_stations.extend(bot_stations)
                logger.info("[TRAIN][EXTRACT] Back on Track : %d stations.", len(bot_stations))
                spark.catalog.clearCache()
            except Exception as e:
                logger.error("[TRAIN][EXTRACT] Erreur Back on Track : %s", e)
                metrics["erreurs"] += 1

            if not all_train_stations:
                logger.error("[TRAIN] Aucune station collectée — section train ignorée.")
                metrics["erreurs"] += 1
            else:
                # ── TRAIN — TRANSFORM ──────────────────────────────────────────
                logger.info("--- TRAIN - TRANSFORM -------------------------------------------")
                canonical_stations, global_mapping = train_build_station_registry(all_train_stations)
                del all_train_stations
                gc.collect()

                # ── TRAIN — LOAD ───────────────────────────────────────────────
                logger.info("--- TRAIN - LOAD ------------------------------------------------")
                # Vider fact_emission juste avant les inserts (train + avion à suivre)
                _run_sql(f"TRUNCATE TABLE {SCHEMA}.fact_emission RESTART IDENTITY;")
                n_stations = train_load_stations(spark, canonical_stations)
                metrics["airports_charges"] += n_stations

                for feed in feeds_with_trains:
                    try:
                        n = train_process_feed(
                            spark, feed, global_mapping,
                            source_id=feed.get("source_id")
                        )
                        metrics["routes_chargees"] += n
                        spark.catalog.clearCache()
                        gc.collect()
                    except Exception as e:
                        logger.error("[TRAIN][LOAD] Erreur feed %s : %s", feed["feed_id"], e)
                        metrics["erreurs"] += 1

                try:
                    n = train_process_backontrack(
                        spark, global_mapping,
                        source_id=source_ids.get("Back on Track")
                    )
                    metrics["routes_chargees"] += n
                    spark.catalog.clearCache()
                    gc.collect()
                except Exception as e:
                    logger.error("[TRAIN][LOAD] Erreur Back on Track : %s", e)
                    metrics["erreurs"] += 1

        # ══════════════════════════════════════════════════════════════════════
        # ETL AVION
        # ══════════════════════════════════════════════════════════════════════
        # Redémarrer Spark pour libérer la mémoire JVM après le section TRAIN
        logger.info("[SPARK] Redémarrage Spark pour la section AVION (libération mémoire JVM)...")
        spark.catalog.clearCache()
        spark.stop()
        import time as _time
        _time.sleep(3)
        gc.collect()
        spark = (
            SparkSession.builder
            .appName("RailCarbon ETL - AVION")
            .config("spark.driver.extraClassPath", _pg_jar)
            .config("spark.executor.extraClassPath", _pg_jar)
            .config("spark.driver.memory", "4g")
            .config("spark.executor.memory", "2g")
            .config("spark.sql.shuffle.partitions", "8")
            .config("spark.python.worker.reuse", "true")
            .config("spark.sql.execution.arrow.pyspark.enabled", "true")
            .getOrCreate()
        )
        logger.info("[SPARK] Session AVION démarrée.")

        logger.info("--- ETL AVION -----------------------------------------------------------")
        # NB: pas de truncate avion  14 dim_route_avion/dim_vehicle_avion utilisent upserts
        # et TRUNCATE CASCADE effacerait fact_emission (train facts) via FK route_avion_id
        # avion_truncate_tables()  <-- SUPPRIME, bug CASCADE

        # ── AVION — EXTRACT ────────────────────────────────────────────────────
        logger.info("--- AVION - EXTRACT -----------------------------------------------------")

        airports_df = avion_load_airports_df(AIRPORTS_CSV, spark)
        if airports_df is None:
            logger.error("[AVION] Index aéroports vide — section avion abandonnée.")
            metrics["erreurs"] += 1
        else:
            # Téléchargement mensuel → cache disque (opdi_cache/)
            # Le cache fait office de checkpoint : si le fichier existe, on le skip
            urls = _avion_generate_urls()
            logger.info("[AVION][EXTRACT] %d mois a traiter (%s -> %s).", len(urls), START_DATE, END_DATE)

            def _dl_month(url):
                month = url.split("_")[-1].replace(".parquet", "")
                return avion_download_to_cache(url, month)

            from concurrent.futures import ThreadPoolExecutor, as_completed
            n_downloaded = 0
            # 4 threads max — les fichiers OPDI sont volumineux (~50-200 Mo chacun)
            with ThreadPoolExecutor(max_workers=4) as executor:
                futures = {executor.submit(_dl_month, u): u for u in urls}
                for fut in as_completed(futures):
                    if fut.result():
                        n_downloaded += 1
                    else:
                        metrics["erreurs"] += 1

            metrics["mois_traites"] = n_downloaded
            logger.info("[AVION][EXTRACT] %d/%d mois disponibles en cache.", n_downloaded, len(urls))

            # Vérifier que le cache contient des fichiers
            cached_files = [
                f for f in os.listdir(OPDI_CACHE_DIR)
                if f.startswith("flight_list_") and f.endswith(".parquet")
            ] if os.path.exists(OPDI_CACHE_DIR) else []

            if not cached_files:
                logger.error("[AVION] Cache vide — section avion abandonnée.")
                metrics["erreurs"] += 1
            else:
                # ── AVION — TRANSFORM ──────────────────────────────────────────
                logger.info("--- AVION - TRANSFORM -------------------------------------------")

                try:
                    routes_df, tc_counts_df = avion_transform_all_flights(spark, airports_df)
                except Exception as _avion_err:
                    logger.error("[AVION][TRANSFORM] Exception : %s", _avion_err, exc_info=True)
                    routes_df, tc_counts_df = None, None
                    metrics["erreurs"] += 1

                if routes_df is None:
                    logger.error("[AVION][TRANSFORM] Echec — section avion abandonnée.")
                    metrics["erreurs"] += 1
                else:
                    final_routes_df = avion_consolidate_routes(routes_df, tc_counts_df)
                    spark.catalog.clearCache()
                    gc.collect()

                    raw_easa  = avion_extract_easa(EASA_FILE, EASA_SHEET)
                    if raw_easa is None:
                        logger.error("[AVION] Impossible de lire EASA — section avion abandonnée.")
                        metrics["erreurs"] += 1
                    else:
                        easa_rows = avion_transform_easa(raw_easa)

                        # ── AVION — LOAD ───────────────────────────────────────
                        logger.info("--- AVION - LOAD ------------------------------------------------")

                        ok_veh = avion_load_vehicle_psql(
                            easa_rows,
                            source_id=source_ids.get("EASA CO2 Database")
                        )
                        if not ok_veh:
                            metrics["erreurs"] += 1

                        n_stations = avion_load_stations_psql(
                            final_routes_df, airports_df,
                            source_id=source_ids.get("OurAirports")
                        )
                        metrics["airports_charges"] += n_stations

                        n_routes = avion_load_routes_psql(
                            final_routes_df,
                            source_id=source_ids.get("OPDI Eurocontrol")
                        )
                        metrics["routes_chargees"] += n_routes

                        avion_update_co2_psql()
                        avion_insert_facts_psql()

                        # Collecte métriques CO2 depuis la BDD
                        try:
                            raw = _query_sql("""
                                SELECT
                                    COUNT(*) FILTER (WHERE dominant_typecode IN (
                                        SELECT icao_typecode FROM mart.dim_vehicle_avion
                                        WHERE icao_typecode IS NOT NULL
                                    )),
                                    COUNT(*) FILTER (WHERE co2_total_kg IS NOT NULL
                                        AND (dominant_typecode IS NULL
                                             OR dominant_typecode NOT IN (
                                                 SELECT icao_typecode FROM mart.dim_vehicle_avion
                                                 WHERE icao_typecode IS NOT NULL
                                             )))
                                FROM mart.dim_route_avion;
                            """)
                            if raw and "|" in raw:
                                parts = raw.split("|")
                                metrics["co2_exact"]    = int(parts[0] or 0)
                                metrics["co2_fallback"] = int(parts[1] or 0)
                        except Exception as e:
                            logger.warning("[AVION][LOAD] Collecte métriques CO2 échouée : %s", e)

    finally:
        spark.catalog.clearCache()
        spark.stop()
        gc.collect()
        logger.info("[SPARK] Session arrêtée.")

    # Finalisation du run log
    statut = "succes" if metrics["erreurs"] == 0 else "erreur"
    run_end(run_id, metrics, statut)

    logger.info("=== Pipeline ETL RailCarbon — terminé ===")
    logger.info(
        "Résumé : mois=%d | stations+aéroports=%d | routes=%d | "
        "co2_exact=%d | co2_fallback=%d | erreurs=%d",
        metrics["mois_traites"],
        metrics["airports_charges"],
        metrics["routes_chargees"],
        metrics["co2_exact"],
        metrics["co2_fallback"],
        metrics["erreurs"],
    )


if __name__ == "__main__":
    main()
